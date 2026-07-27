"""
planilla.import.data.wizard
Importacion masiva desde el machote Excel generado por import_template_wizard.
Estrategia: si el empleado ya existe (por identification_id) -> se salta.
Al finalizar: resumen en pantalla + Excel de errores descargable.
"""
from odoo import models, fields, api
from ..models import planilla_const as K
from odoo.exceptions import UserError
import base64, io, logging, re, traceback
from datetime import date, datetime

_logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ==============================================================================
# TABLAS DE TRADUCCION  (valor amigable del Excel -> valor tecnico del modelo)
# ==============================================================================

# Las 18 constantes de mapeo (INS_RISK, INS_WORKDAY, INS_NATIONALITY,
# ACCOUNT_TYPE, GENDER, INS_CIVIL, INS_ID_TYPE, INS_ID_TYPE_CODE,
# DISABILITY_TYPE, BENEFIT_TYPE, AMOUNT_TYPE, PENSION_CALC,
# OVERTIME_TYPE, BANK, CALC_METHOD, LOAN_TYPE, LOAN_STATE,
# PENSION_RELATION) que vivian aqui se centralizaron en
# planilla_const.py (K.NOMBRE_CONSTANTE). Ninguna se usaba dentro de
# este archivo -- solo se definian aqui y los procesadores en
# wizard/processors/*.py las necesitaban sin poder importarlas, lo que
# causaba NameError en cada importacion masiva por Excel de empleados,
# prestamos, pensiones, incapacidades y horas extra.


# FREQUENCY: mapeo de texto a codigo de frecuencia -- centralizado en
# planilla_const.py (K.FREQUENCY). Se elimino la copia local que vivia
# aqui porque nunca se usaba dentro de este archivo, y su existencia
# aislada fue la causa de que wizard/processors/proc_employees.py
# lanzara NameError al no poder encontrarla (ese archivo no la importaba
# de aqui, y no existia en ningun otro lado accesible para el).


from .import_parse_utils import (
    BOOL_MAP, _normalize, _parse_date, _parse_float, _parse_int, _parse_bool, _map,
)


# ==============================================================================
# MODELO WIZARD
# ==============================================================================

class ImportDataWizard(models.TransientModel):
    _name        = 'planilla.import.data.wizard'
    _description = 'Importacion Masiva de Empleados desde Excel'
    _inherit = [
        'planilla.import.processor.employees',
        'planilla.import.processor.loans',
        'planilla.import.processor.novedades',
        'planilla.import.processor.bonos.embargos',
    ]

    company_id = fields.Many2one(
        'res.company', required=True, default=lambda self: self.env.company)

    excel_file = fields.Binary(
        string='Archivo Excel (Machote)', required=True,
        help='Cargue el machote Excel completado por el cliente.')
    excel_filename = fields.Char(string='Nombre de archivo')

    # Hojas a procesar
    import_employees    = fields.Boolean('  Empleados',                      default=True)
    import_loans        = fields.Boolean('  Prestamos',                      default=True)
    import_pension      = fields.Boolean('  Pensiones Alimentarias',         default=True)
    import_benefits     = fields.Boolean('  Beneficios / Deducciones',       default=True)
    import_disabilities = fields.Boolean('  Incapacidades',                  default=True)
    import_vacations    = fields.Boolean('  Vacaciones',                     default=True)
    import_overtime     = fields.Boolean('  Horas Extras',                   default=True)
    import_embargos     = fields.Boolean('  Embargos Judiciales',            default=True)
    import_acumulados   = fields.Boolean('  Acumulados (aguinaldo/vacaciones)', default=True)
    import_bonos        = fields.Boolean('  Bonos y Beneficios',             default=True)
    import_sample_data  = fields.Boolean(
        '  Importar fila de prueba (cedula 1-0000-0001)',
        default=False,
        help='Active solo cuando quiera verificar que la importacion funciona. '
             'Desactivado por defecto para que la fila naranja se ignore automaticamente.'
    )
    sample_exists = fields.Boolean(
        string='Empleado de prueba existe',
        compute='_compute_sample_exists',
    )

    @api.depends('company_id')
    def _compute_sample_exists(self):
        for rec in self:
            rec.sample_exists = bool(
                self.env['hr.employee'].search([
                    ('identification_id', '=', rec._SAMPLE_CEDULA),
                    ('company_id', '=', rec.company_id.id),
                ], limit=1)
            )

    # FIX v512 BUG-07: unificada con K.TEST_CEDULA (era '1-0000-0001' hardcoded en la clase,
    # desincronizado de planilla_const.py). Ahora un solo punto de verdad.
    _SAMPLE_CEDULA = K.TEST_CEDULA

    # Resultados (readonly, visibles despues de procesar)
    state = fields.Selection([
        ('draft',  'Pendiente'),
        ('done',   'Procesado'),
    ], default='draft')

    # Contadores resumen
    emp_created  = fields.Integer('Empleados creados',     readonly=True)
    emp_skipped  = fields.Integer('Empleados existentes (omitidos)', readonly=True)
    emp_errors   = fields.Integer('Empleados con errores', readonly=True)
    loan_created = fields.Integer('Prestamos creados',     readonly=True)
    pen_created  = fields.Integer('Pensiones creadas',     readonly=True)
    ben_created  = fields.Integer('Beneficios creados',    readonly=True)
    dis_created  = fields.Integer('Incapacidades creadas', readonly=True)
    vac_created  = fields.Integer('Saldos de vacaciones procesados', readonly=True)
    ot_created   = fields.Integer('Horas extras creadas',  readonly=True)
    emb_created  = fields.Integer('Embargos creados',      readonly=True)
    bon_created  = fields.Integer('Bonos creados',         readonly=True)
    total_errors = fields.Integer('Total errores',         readonly=True)

    result_summary = fields.Text('Resumen',   readonly=True)
    report_file    = fields.Binary('Reporte de Errores (Excel)', readonly=True)
    report_name    = fields.Char(default='Reporte_Importacion.xlsx')

    # -- helpers internos ------------------------------------------------------

    def _get_wb(self):
        if not OPENPYXL_OK:
            raise UserError('La libreria openpyxl no esta instalada en el servidor.')
        raw = base64.b64decode(self.excel_file)
        return load_workbook(io.BytesIO(raw), data_only=True, read_only=True)

    def _sheet_rows(self, wb, sheet_names):
        """Retorna (headers_dict, filas) de la primera hoja encontrada."""
        _EXAMPLE_CEDULA = K.SAMPLE_CEDULA   # fila verde de ejemplo -- siempre se salta
        for name in sheet_names:
            for sn in wb.sheetnames:
                if name.lower() in sn.lower():
                    ws = wb[sn]
                    rows = list(ws.iter_rows(values_only=True))
                    if len(rows) < 2:
                        return {}, []

                    # -- Detectar fila de encabezados ----------------------
                    # Buscamos la fila donde alguna celda sea EXACTAMENTE una
                    # palabra clave de encabezado (no un titulo largo que las
                    # contenga como substring). Esto distingue la fila de
                    # encabezados de columna de la fila de titulo o secciones.
                    EXACT_KEYS = ('nombre completo', 'cedula / identificacion',
                                  'cedula empleado', 'numero de expediente',
                                  'concepto', 'tipo de incapacidad',
                                  'dias acumulados', 'tipo de hora extra',
                                  'tipo de prestamo', 'cedula empleado')
                    PARTIAL_KEYS = ('cedula', 'cedula')
                    header_row = None
                    for i, r in enumerate(rows):
                        cells = [str(c).strip().lower() for c in r
                                 if c is not None and str(c).strip()]
                        # Coincidencia exacta con alguna clave conocida
                        if any(cell in EXACT_KEYS for cell in cells):
                            header_row = i
                            break
                        # O celda que sea exactamente 'cedula' o 'cedula'
                        if any(cell in PARTIAL_KEYS for cell in cells):
                            header_row = i
                            break

                    if header_row is None:
                        return {}, []

                    hdrs = {str(c).strip(): ci
                            for ci, c in enumerate(rows[header_row]) if c}

                    # -- Datos: todo lo que viene despues del encabezado ---
                    data_rows = rows[header_row + 1:]

                    # Saltar la fila de ejemplo verde (cedula 1-2345-6789)
                    # buscando ese valor en cualquier columna de cada fila
                    data_rows = [
                        r for r in data_rows
                        if not any(
                            str(c).strip() == _EXAMPLE_CEDULA
                            for c in r if c is not None
                        )
                    ]

                    # Filtrar filas completamente vacias
                    data_rows = [r for r in data_rows
                                 if any(c for c in r
                                        if c is not None and str(c).strip())]

                    # FIX v5.15.6: Filtrar filas de instrucciones al pie de la hoja.
                    # Algunos machotes tienen un bloque de instrucciones al final
                    # (ej. fila 85+ en VACACIONES) que el wizard leia como datos.
                    # Una fila es instruccion si su primera celda no-nula:
                    #   - empieza con '' (marcador de instrucciones), o
                    #   - tiene mas de 80 chars (ninguna cedula/nombre es tan largo), o
                    #   - empieza con un numero seguido de punto y espacio ('1. ', '2. ')
                    def _is_instruction_row(row):
                        import re as _re
                        INSTR_KEYWORDS = (
                            'instruccion', 'instruction', 'nota:', 'note:',
                            'como calcular', 'ejemplo:', 'example:',
                        )
                        for c in row:
                            if c is not None and str(c).strip():
                                txt = str(c).strip()
                                txt_l = txt.lower()
                                # Fila de prueba WARN
                                if txt_l.startswith('warn ') or txt_l.startswith('warn:'):
                                    return True
                                # Instruccion indentada (doble espacio)
                                if txt.startswith('  '):
                                    return True
                                # Texto muy largo = instruccion
                                if len(txt) > 100:
                                    return True
                                # Instruccion numerada: "1. ", "2. " etc.
                                if _re.match(r'^\d+\.\s+', txt):
                                    return True
                                # Palabras clave de instrucciones
                                if any(kw in txt_l for kw in INSTR_KEYWORDS):
                                    return True
                                break
                        return False

                    # Truncar en el primer bloque de instrucciones encontrado
                    clean_rows = []
                    for r in data_rows:
                        if _is_instruction_row(r):
                            break  # todo lo que sigue es instrucciones
                        clean_rows.append(r)
                    data_rows = clean_rows

                    return hdrs, data_rows
        return {}, []

    def _v(self, row, hdrs, *col_names):
        """Obtiene valor de la primera columna que coincida."""
        for name in col_names:
            for hdr, idx in hdrs.items():
                if name.lower() in hdr.lower():
                    val = row[idx] if idx < len(row) else None
                    if val is not None and str(val).strip():
                        return val
        return None

    def _find_employee(self, cedula):
        """Busca empleado por identification_id."""
        if not cedula:
            return None
        cedula = str(cedula).strip()
        return self.env['hr.employee'].search(
            [('identification_id', '=', cedula),
             ('company_id', '=', self.company_id.id)], limit=1)

    def _find_m2o(self, model, name_val, field='name', extra_domain=None):
        """Busca un registro por nombre (case-insensitive).
        Usa sudo() + active_test=False para bypassear ir.rules y filtros
        de registros archivados. Los registros pueden tener company_id=NULL
        o pertenecer a otra empresa en la misma BD.
        """
        if not name_val:
            return None
        name_str = str(name_val).strip()
        domain = [(field, 'ilike', name_str)]
        if extra_domain:
            domain += extra_domain
        # FIX PEND-04: agregar filtro de empresa para evitar cross-company
        # en entornos multi-empresa. Solo aplica si el modelo tiene company_id.
        try:
            if 'company_id' in self.env[model]._fields:
                domain += ['|',
                    ('company_id', '=', self.company_id.id),
                    ('company_id', '=', False)]
        except Exception:
            pass
        result = self.env[model].sudo().with_context(active_test=False).search(
            domain, limit=1)
        return result or None

    # ==========================================================================
    # PROCESADORES POR HOJA
    # ==========================================================================

    def _is_sample(self, cedula):
        """Retorna True si la cedula corresponde a la fila de prueba."""
        return str(cedula).strip() == self._SAMPLE_CEDULA








    def _process_acumulados(self, wb, errors):
        """
        Procesa la hoja ACUMULADOS: carga el acumulado de aguinaldo inicial
        y el saldo de vacaciones en dias directamente en el empleado.

        Columnas esperadas:
          - Cedula Empleado       (obligatoria)
          - Fecha de Corte        (DD/MM/AAAA)
          - Aguinaldo Acumulado   (CRC) -> aguinaldo_initial_amount
          - Vacaciones Acumuladas (dias) -> vacation_initial_balance

        Si el empleado ya tiene vacaciones cargadas desde la hoja VACACIONES,
        no sobreescribe (la hoja VACACIONES tiene prioridad).
        """
        created = err_count = 0
        hdrs, rows = self._sheet_rows(wb, ['ACUMULADO', 'ACUMULADOS', 'PROVISION'])
        if not rows:
            return 0, 0

        for row_num, row in enumerate(rows, start=1):
            v = lambda *cols: self._v(row, hdrs, *cols)
            cedula = str(v('Cedula', 'Cedula Empleado') or '').strip()
            if not cedula:
                continue
            if self._is_sample(cedula) and not self.import_sample_data:
                continue

            emp = self._find_employee(cedula)
            if not emp:
                err_count += 1
                errors.append({
                    'hoja': 'ACUMULADOS', 'fila': row_num, 'cedula': cedula,
                    'nombre': '', 'error': 'Empleado no encontrado en Odoo',
                })
                continue

            try:
                with self.env.cr.savepoint():
                    fecha_corte = _parse_date(
                        v('Fecha de Corte', 'Fecha Corte', 'Fecha')
                    )
                    ag_raw  = v('Aguinaldo Acumulado', 'Aguinaldo Acumulado (CRC)',
                                'Aguinaldo', 'Acumulado Aguinaldo')
                    vac_raw = v('Vacaciones Acumuladas', 'Vacaciones Acumuladas (dias)',
                                'Vacaciones', 'Dias Vacaciones')

                    ag_amount = _parse_float(ag_raw)
                    vac_dias  = _parse_float(vac_raw)

                    vals_emp = {}

                    # Aguinaldo inicial
                    if ag_amount and ag_amount > 0:
                        vals_emp['aguinaldo_initial_amount'] = round(ag_amount, 2)
                        if fecha_corte:
                            vals_emp['aguinaldo_initial_date'] = fecha_corte

                    # Vacaciones iniciales (solo si NO tiene ya vacation_initial_balance_date)
                    if vac_dias and not emp.vacation_initial_balance_date:
                        vals_emp['vacation_initial_balance'] = round(vac_dias, 2)
                        if fecha_corte:
                            vals_emp['vacation_initial_balance_date'] = fecha_corte

                    if vals_emp:
                        emp.with_context(skip_salary_history=True).write(vals_emp)
                        if 'vacation_initial_balance' in vals_emp:
                            emp._compute_vacation_balance()
                        created += 1

            except Exception as e:
                import traceback
                err_count += 1
                errors.append({
                    'hoja': 'ACUMULADOS', 'fila': row_num, 'cedula': cedula,
                    'nombre': emp.name if emp else '',
                    'error': str(e)[:200],
                    'traceback': traceback.format_exc(),
                    'vals': {},
                })
                _logger.warning('ImportDataWizard ACUMULADOS fila %s: %s', row_num, e)

        return created, err_count


    # ==========================================================================
    # PROCESADOR EMBARGOS JUDICIALES
    # ==========================================================================


    # ==========================================================================
    # PROCESADOR BONOS E INCENTIVOS
    # ==========================================================================


    # ==========================================================================
    # REPORTE DE ERRORES EN EXCEL
    # ==========================================================================

    def _build_error_report(self, errors, counters):
        wb = Workbook()

        DARK  = '1F3864'
        GREEN = '375623'
        RED   = 'C00000'
        AMBER = 'BF8F00'
        LGRAY = 'F2F2F2'
        ORANGE= 'F4B942'
        WHITE = 'FFFFFF'
        BLUE  = '2E75B6'

        def fill(c):  return PatternFill('solid', fgColor=c)
        def font(bold=False, color='000000', size=10, italic=False):
            return Font(bold=bold, color=color, size=size, italic=italic, name='Arial')
        def bdr():
            s = Side(style='thin', color='BDD7EE')
            return Border(left=s, right=s, top=s, bottom=s)
        def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
        def left(wrap=True): return Alignment(horizontal='left', vertical='top', wrap_text=wrap)

        total_errors = sum(e for e in [
            counters['emp_errors'], counters['loan_errors'], counters['pen_errors'],
            counters['ben_errors'], counters['dis_errors'], counters['vac_errors'],
            counters['ot_errors'],
            # FIX-O4: faltaban embargos y bonos -- el resumen subestimaba errores
            counters.get('emb_errors', 0), counters.get('bon_errors', 0),
        ] if e)
        # FIX ALERTA-01: las advertencias (campos que no se pudieron vincular,
        # IBAN invalido, etc.) se guardan en la misma lista `errors` con el
        # prefijo 'ADVERTENCIA:' pero NO incrementan err_count -- el registro
        # SI se creo correctamente, solo con datos incompletos. Antes esto
        # significaba que el usuario nunca se enteraba: el resumen decia
        # "SIN ERRORES" aunque hubiera advertencias reales en el detalle.
        total_warnings = sum(
            1 for e in errors if str(e.get('error', '')).startswith('ADVERTENCIA')
        )
        total_ok = sum(e for e in [
            counters['emp_created'], counters['loan_created'], counters['pen_created'],
            counters['ben_created'], counters['dis_created'], counters['vac_created'],
            counters['ot_created'],
            # FIX-O4: faltaban embargos y bonos -- el resumen subestimaba importaciones
            counters.get('emb_created', 0), counters.get('bon_created', 0),
        ] if e)

        # ======================================================================
        # HOJA 1 -- RESUMEN
        # ======================================================================
        ws1 = wb.active
        ws1.title = ' Resumen'
        ws1.sheet_view.showGridLines = False

        # Cabecera
        ws1.merge_cells('A1:G1')
        c = ws1['A1']
        c.value = f'REPORTE DE IMPORTACION -- {self.company_id.name}'
        c.font = font(bold=True, color=WHITE, size=14)
        c.fill = fill(DARK)
        c.alignment = center()
        ws1.row_dimensions[1].height = 34

        ws1.merge_cells('A2:G2')
        c = ws1['A2']
        c.value = (f'Archivo: {self.excel_filename or "--"}   |   '
                   f'Procesado: {datetime.now().strftime("%d/%m/%Y %H:%M")}   |   '
                   f'Total creados: {total_ok}   |   Total errores: {total_errors}'
                   f'{f"   |   Advertencias: {total_warnings}" if total_warnings else ""}')
        c.font = font(italic=True, color=WHITE, size=9)
        c.fill = fill(BLUE)
        c.alignment = center()
        ws1.row_dimensions[2].height = 16

        # Resultado global
        ws1.row_dimensions[3].height = 8
        ws1.merge_cells('A4:G4')
        if total_errors:
            estado_txt = f'WARN  IMPORTACION CON {total_errors} ERROR(ES) -- Ver hoja "Detalle Errores"'
        elif total_warnings:
            estado_txt = (
                f'AVISO IMPORTACION COMPLETADA CON {total_warnings} ADVERTENCIA(S) '
                f'-- Algunos registros se crearon con campos incompletos. '
                f'Ver hoja "Detalle Errores"'
            )
        else:
            estado_txt = 'OK IMPORTACION COMPLETADA SIN ERRORES'
        c = ws1['A4']
        c.value = estado_txt
        c.font = font(bold=True, color=WHITE, size=11)
        if total_errors:
            c.fill = fill(RED)
        elif total_warnings:
            c.fill = fill(AMBER)
        else:
            c.fill = fill('375623')
        c.alignment = center()
        ws1.row_dimensions[4].height = 26

        # Tabla resumen por hoja
        ws1.row_dimensions[5].height = 8
        hdrs_res = ['Hoja', 'Registros Creados OK', 'Omitidos ', 'Errores ERR', 'Total Procesados']
        for ci, h in enumerate(hdrs_res, 1):
            c = ws1.cell(6, ci, value=h)
            c.font = font(bold=True, color=WHITE, size=10)
            c.fill = fill(BLUE)
            c.border = bdr()
            c.alignment = center()
        ws1.row_dimensions[6].height = 22

        summary_data = [
            (' Empleados',      counters['emp_created'],  counters['emp_skipped'],  counters['emp_errors']),
            (' Prestamos',      counters['loan_created'], 0,                        counters['loan_errors']),
            (' Pensiones',       counters['pen_created'],  0,                        counters['pen_errors']),
            (' Otros Descuentos',counters['ben_created'],  0,                        counters['ben_errors']),
            (' Incapacidades',   counters['dis_created'],  0,                        counters['dis_errors']),
            (' Vacaciones',      counters['vac_created'],  0,                        counters['vac_errors']),
            (' Horas Extras',    counters['ot_created'],   0,                        counters['ot_errors']),
            (' Embargos',        counters.get('emb_created', 0), 0,                 counters.get('emb_errors', 0)),
            (' Bonos',           counters.get('bon_created', 0), 0,                 counters.get('bon_errors', 0)),
        ]

        for i, (hoja, created, skipped, errs) in enumerate(summary_data, start=7):
            total = created + skipped + errs
            row_data = [hoja, created, skipped, errs, total]
            for ci, val in enumerate(row_data, 1):
                c = ws1.cell(i, ci, value=val)
                c.border = bdr()
                c.alignment = center() if ci > 1 else left(wrap=False)
                if ci == 1:
                    c.fill = fill('EBF3FB')
                    c.font = font(bold=True, size=10)
                elif ci == 2:
                    c.fill = fill('E2EFDA' if created else WHITE)
                    c.font = font(bold=bool(created), color=GREEN if created else '888888', size=10)
                elif ci == 3:
                    c.fill = fill('FFF9E6' if skipped else WHITE)
                    c.font = font(size=10, color=AMBER if skipped else '888888')
                elif ci == 4:
                    c.fill = fill('FCE4D6' if errs else WHITE)
                    c.font = font(bold=bool(errs), color=RED if errs else '888888', size=10)
                else:
                    c.fill = fill(LGRAY)
                    c.font = font(size=10)
            ws1.row_dimensions[i].height = 18

        # Totales
        tot_row = 7 + len(summary_data)
        tot_data = ['TOTAL', total_ok,
                    counters['emp_skipped'],
                    total_errors,
                    total_ok + counters['emp_skipped'] + total_errors]
        for ci, val in enumerate(tot_data, 1):
            c = ws1.cell(tot_row, ci, value=val)
            c.font = font(bold=True, size=10, color=WHITE)
            c.fill = fill(DARK)
            c.border = bdr()
            c.alignment = center() if ci > 1 else left(wrap=False)
        ws1.row_dimensions[tot_row].height = 20

        # Anchos hoja 1
        for col, w in [(1,26),(2,22),(3,16),(4,14),(5,18)]:
            ws1.column_dimensions[get_column_letter(col)].width = w

        # ======================================================================
        # HOJA 2 -- DETALLE DE ERRORES
        # ======================================================================
        ws2 = wb.create_sheet('ERR Detalle Errores')
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells('A1:H1')
        c = ws2['A1']
        c.value = f'DETALLE DE ERRORES -- {self.company_id.name}  |  {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        c.font = font(bold=True, color=WHITE, size=12)
        c.fill = fill(RED)
        c.alignment = center()
        ws2.row_dimensions[1].height = 28

        if not errors:
            ws2.merge_cells('A2:H2')
            c = ws2['A2']
            c.value = 'OK No hubo errores en esta importacion'
            c.font = font(bold=True, color=GREEN, size=12)
            c.fill = fill('E2EFDA')
            c.alignment = center()
            ws2.row_dimensions[2].height = 28
        else:
            hdrs2 = ['#', 'Hoja', 'Fila Excel', 'Cedula', 'Nombre', 'Tipo de Error', 'Mensaje de Error', 'Accion Sugerida']
            for ci, h in enumerate(hdrs2, 1):
                c = ws2.cell(2, ci, value=h)
                c.font = font(bold=True, color=WHITE, size=9)
                c.fill = fill(DARK)
                c.border = bdr()
                c.alignment = center()
            ws2.row_dimensions[2].height = 20

            def _classify_error(err_msg):
                """Clasifica el error y sugiere accion correctiva."""
                msg = str(err_msg).lower()
                if str(err_msg).startswith('ADVERTENCIA'):
                    return 'Advertencia (no bloqueante)', (
                        'El registro SI se creo. Revise y complete el/los '
                        'campo(s) mencionados manualmente en Odoo si son '
                        'necesarios.'
                    )
                if 'not found' in msg or 'no encontrado' in msg:
                    return 'Empleado no existe', 'Importe primero la hoja EMPLEADOS antes de importar otras hojas'
                if 'wrong value' in msg or 'invalid value' in msg:
                    campo = ''
                    import re
                    m = re.search(r"'([^']+)':\s*'([^']+)'", err_msg)
                    if m:
                        campo = m.group(1).split('.')[-1]
                        valor = m.group(2)
                        return f'Valor invalido en {campo}', f'El valor "{valor}" no es valido para el campo {campo}. Consulte la hoja CATALOGOS'
                    return 'Valor invalido', 'Verifique que el valor corresponda exactamente a los listados en la hoja CATALOGOS'
                if 'invalid field' in msg:
                    import re
                    m = re.search(r"'([^']+)'", err_msg)
                    campo = m.group(1) if m else ''
                    return f'Campo inexistente: {campo}', f'El campo {campo} no existe en esta version de Odoo. Contacte al administrador'
                if 'required' in msg or 'obligatorio' in msg or 'cannot be empty' in msg:
                    return 'Campo obligatorio vacio', 'Complete todos los campos marcados en amarillo en el machote'
                if 'unique' in msg or 'duplicate' in msg or 'ya existe' in msg:
                    return 'Registro duplicado', 'Este registro ya existe en Odoo. Verifique la cedula y elimine el duplicado'
                if 'constraint' in msg:
                    return 'Restriccion de BD', 'Violacion de regla de negocio. Revise los datos ingresados'
                if 'date' in msg or 'fecha' in msg:
                    return 'Error de fecha', 'Use el formato DD/MM/AAAA. Ejemplo: 01/03/2024'
                if 'float' in msg or 'int' in msg or 'numeric' in msg:
                    return 'Error numerico', 'Use solo numeros sin comas ni simbolos. Ejemplo: 750000'
                return 'Error inesperado', 'Revise el log tecnico en la hoja "Log Tecnico" para mas detalles'

            for num, err in enumerate(errors, start=1):
                r = 2 + num
                err_msg = err.get('error', '')
                tipo, accion = _classify_error(err_msg)
                row_data = [
                    num,
                    err.get('hoja', ''),
                    err.get('fila', ''),
                    err.get('cedula', ''),
                    err.get('nombre', ''),
                    tipo,
                    err_msg,
                    accion,
                ]
                is_odd = num % 2 == 0
                row_bg = 'FFF9F9' if is_odd else WHITE
                is_warning = str(err_msg).startswith('ADVERTENCIA')
                for ci, val in enumerate(row_data, 1):
                    c = ws2.cell(r, ci, value=val)
                    c.border = bdr()
                    c.alignment = left() if ci >= 6 else center()
                    if ci == 7:  # mensaje error
                        if is_warning:
                            c.font = font(size=9, color=AMBER, bold=True)
                            c.fill = fill('FFF3D6')
                        else:
                            c.font = font(size=9, color=RED, bold=True)
                            c.fill = fill('FCE4D6')
                    elif ci == 8:  # accion
                        c.font = font(size=9, color='7B3F00', italic=True)
                        c.fill = fill('FFF2CC')
                    elif ci == 6:  # tipo
                        c.font = font(size=9, color=DARK, bold=True)
                        c.fill = fill('EBF3FB')
                    else:
                        c.font = font(size=9)
                        c.fill = fill(row_bg)
                ws2.row_dimensions[r].height = 40

        for col, w in [(1,5),(2,18),(3,10),(4,16),(5,22),(6,24),(7,48),(8,46)]:
            ws2.column_dimensions[get_column_letter(col)].width = w

        # ======================================================================
        # HOJA 3 -- LOG TECNICO
        # ======================================================================
        ws3 = wb.create_sheet(' Log Tecnico')
        ws3.sheet_view.showGridLines = False

        ws3.merge_cells('A1:C1')
        c = ws3['A1']
        c.value = f'LOG TECNICO -- {self.company_id.name}  |  {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        c.font = font(bold=True, color=WHITE, size=11)
        c.fill = fill('4A4A4A')
        c.alignment = center()
        ws3.row_dimensions[1].height = 26

        ws3.merge_cells('A2:C2')
        c = ws3['A2']
        c.value = 'Este log contiene el traceback completo y los valores enviados en cada error. Comparta esta hoja con soporte tecnico.'
        c.font = font(italic=True, size=9, color='555555')
        c.fill = fill('F5F5F5')
        c.alignment = left()
        ws3.row_dimensions[2].height = 16

        log_row = 4
        if not errors:
            ws3.merge_cells('A3:C3')
            c = ws3['A3']
            c.value = 'OK Sin errores -- no hay log tecnico que mostrar'
            c.font = font(bold=True, color=GREEN, size=10)
            c.fill = fill('E2EFDA')
            c.alignment = center()
        else:
            for num, err in enumerate(errors, start=1):
                # Bloque por error
                ws3.merge_cells(f'A{log_row}:C{log_row}')
                c = ws3.cell(log_row, 1,
                    value=f'ERROR #{num} -- Hoja: {err.get("hoja","")}  |  Fila: {err.get("fila","")}  |  Cedula: {err.get("cedula","")}  |  Nombre: {err.get("nombre","")}')
                c.font = font(bold=True, color=WHITE, size=10)
                c.fill = fill(RED if num % 2 else '8B0000')
                c.alignment = left(wrap=False)
                ws3.row_dimensions[log_row].height = 20
                log_row += 1

                # Mensaje de error
                ws3.cell(log_row, 1, value='MENSAJE:').font = font(bold=True, size=9, color=DARK)
                ws3.cell(log_row, 1).fill = fill('EBF3FB')
                ws3.merge_cells(f'B{log_row}:C{log_row}')
                c = ws3.cell(log_row, 2, value=err.get('error', ''))
                c.font = font(size=9, color=RED, bold=True)
                c.fill = fill('FCE4D6')
                c.alignment = left()
                ws3.row_dimensions[log_row].height = 30
                log_row += 1

                # Valores enviados
                vals = err.get('vals', {})
                if vals:
                    ws3.cell(log_row, 1, value='VALORES\nENVIADOS:').font = font(bold=True, size=9, color=DARK)
                    ws3.cell(log_row, 1).fill = fill('EBF3FB')
                    ws3.cell(log_row, 1).alignment = center()
                    ws3.merge_cells(f'B{log_row}:C{log_row}')
                    vals_txt = '\n'.join(f'  {k}: {v}' for k, v in vals.items() if v and v != 'False')
                    c = ws3.cell(log_row, 2, value=vals_txt or '(sin valores)')
                    c.font = font(size=8, color='333333')
                    c.fill = fill('F8F8F8')
                    c.alignment = left()
                    ws3.row_dimensions[log_row].height = max(14 * max(1, len(vals_txt.split('\n'))), 30)
                    log_row += 1

                # Traceback
                tb = err.get('traceback', '')
                if tb and tb.strip():
                    ws3.cell(log_row, 1, value='TRACEBACK:').font = font(bold=True, size=9, color=DARK)
                    ws3.cell(log_row, 1).fill = fill('EBF3FB')
                    ws3.cell(log_row, 1).alignment = center()
                    ws3.merge_cells(f'B{log_row}:C{log_row}')
                    c = ws3.cell(log_row, 2, value=tb.strip())
                    c.font = Font(name='Courier New', size=8, color='444444')
                    c.fill = fill('1E1E1E' if False else 'F0F0F0')
                    c.alignment = left()
                    lines = tb.count('\n')
                    ws3.row_dimensions[log_row].height = max(14 * min(lines, 20), 40)
                    log_row += 1

                # Separador
                log_row += 1

        ws3.column_dimensions['A'].width = 16
        ws3.column_dimensions['B'].width = 60
        ws3.column_dimensions['C'].width = 60

        # -- Serializar --------------------------------------------------------
        buf = io.BytesIO()
        wb.save(buf)
        return base64.b64encode(buf.getvalue())

    # ==========================================================================
    # ACCION PRINCIPAL
    # ==========================================================================

    def action_import(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError('Debe cargar el archivo Excel.')

        wb     = self._get_wb()
        errors = []

        # -- Procesar hojas ----------------------------------------------------
        emp_c = emp_s = emp_e = 0
        loan_c = loan_e = 0
        pen_c = pen_e = 0
        ben_c = ben_e = 0
        dis_c = dis_e = 0
        vac_c = vac_e = 0
        ot_c  = ot_e  = 0
        emb_c = emb_e = 0
        bon_c = bon_e = 0

        if self.import_employees:
            emp_c, emp_s, emp_e = self._process_employees(wb, errors)
        if self.import_loans:
            loan_c, loan_e = self._process_loans(wb, errors)
        if self.import_pension:
            pen_c, pen_e = self._process_pension(wb, errors)
        if self.import_benefits:
            ben_c, ben_e = self._process_benefits(wb, errors)
        if self.import_disabilities:
            dis_c, dis_e = self._process_disabilities(wb, errors)
        if self.import_vacations:
            vac_c, vac_e = self._process_vacations(wb, errors)
        acum_c = acum_e = 0
        if self.import_acumulados:
            acum_c, acum_e = self._process_acumulados(wb, errors)
        if self.import_overtime:
            ot_c, ot_e = self._process_overtime(wb, errors)
        if self.import_embargos:
            emb_c, emb_e = self._process_embargos(wb, errors)
        if self.import_bonos:
            bon_c, bon_e = self._process_bonos(wb, errors)

        total_err = emp_e + loan_e + pen_e + ben_e + dis_e + vac_e + acum_e + ot_e + emb_e + bon_e

        counters = {
            'emp_created': emp_c, 'emp_skipped': emp_s, 'emp_errors': emp_e,
            'loan_created': loan_c, 'loan_errors': loan_e,
            'pen_created':  pen_c,  'pen_errors':  pen_e,
            'ben_created':  ben_c,  'ben_errors':  ben_e,
            'dis_created':  dis_c,  'dis_errors':  dis_e,
            'vac_created':  vac_c,  'vac_errors':  vac_e,
            'acum_created': acum_c, 'acum_errors': acum_e,
            'ot_created':   ot_c,   'ot_errors':   ot_e,
            'emb_created':  emb_c,  'emb_errors':  emb_e,
            'bon_created':  bon_c,  'bon_errors':  bon_e,
        }

        # -- Resumen texto -----------------------------------------------------
        summary = (
            f"== IMPORTACION COMPLETADA ==\n\n"
            f" Empleados:        {emp_c} creados  |  {emp_s} omitidos (ya existian)  |  {emp_e} errores\n"
            f" Prestamos:        {loan_c} creados  |  {loan_e} errores\n"
            f" Pensiones:         {pen_c} creadas  |  {pen_e} errores\n"
            f" Otros Descuentos:  {ben_c} creados  |  {ben_e} errores\n"
            f" Incapacidades:    {dis_c} creadas  |  {dis_e} errores\n"
            f" Vacaciones:       {vac_c} procesadas  |  {vac_e} errores\n"
            f" Horas Extras:     {ot_c} creadas  |  {ot_e} errores\n"
            f" Embargos:         {emb_c} creados  |  {emb_e} errores\n"
            f" Bonos:            {bon_c} creados  |  {bon_e} errores\n"
            f"\n{'WARN  Hay errores -- descargue el reporte para ver el detalle.' if total_err else 'OK Sin errores.'}"
        )

        # -- Generar reporte Excel ---------------------------------------------
        report_data = self._build_error_report(errors, counters)
        fname = f'Reporte_Importacion_{self.company_id.name}_{date.today()}.xlsx'

        self.write({
            'state':          'done',
            'emp_created':    emp_c, 'emp_skipped':  emp_s, 'emp_errors':  emp_e,
            'loan_created':   loan_c,
            'pen_created':    pen_c,
            'ben_created':    ben_c,
            'dis_created':    dis_c,
            'vac_created':    vac_c,
            'ot_created':     ot_c,
            'emb_created':    emb_c,
            'bon_created':    bon_c,
            'total_errors':   total_err,
            'result_summary': summary,
            'report_file':    report_data,
            'report_name':    fname,
        })

        # -- Reabrir wizard para mostrar resultados ----------------------------
        return {
            'type':    'ir.actions.act_window',
            'res_model': self._name,
            'res_id':  self.id,
            'view_mode': 'form',
            'view_id': self.env.ref('planilla_cr.view_import_data_wizard_form').id,
            'target': 'new',
        }

    def action_delete_sample(self):
        """Elimina el EMPLEADO PRUEBA y todos sus registros relacionados."""
        self.ensure_one()
        emp = self.env['hr.employee'].search([
            ('identification_id', '=', self._SAMPLE_CEDULA),
            ('company_id', '=', self.company_id.id),
        ], limit=1)

        if not emp:
            raise UserError('No se encontro ningun empleado de prueba (cedula 1-0000-0001) '
                            'en esta empresa. Es posible que ya haya sido eliminado.')

        deleted = []

        # Eliminar registros relacionados antes del empleado
        related = [
            ('planilla.employee.loan',      [('employee_id', '=', emp.id)], 'prestamos'),
            ('planilla.pension.alimentaria',[('employee_id', '=', emp.id)], 'pensiones'),
            ('planilla.recurring.benefit',  [('employee_id', '=', emp.id)], 'beneficios'),
            ('planilla.disability',         [('employee_id', '=', emp.id)], 'incapacidades'),
            ('planilla.vacation.payment',   [('employee_id', '=', emp.id)], 'vacaciones'),
            ('planilla.overtime',           [('employee_id', '=', emp.id)], 'horas extras'),
            ('planilla.embargo',            [('employee_id', '=', emp.id)], 'embargos'),
            ('planilla.bono',               [('employee_id', '=', emp.id)], 'bonos'),
            ('planilla.payslip.cr',         [('employee_id', '=', emp.id)], 'boletas'),
            ('planilla.salary.history',     [('employee_id', '=', emp.id)], 'historial salarial'),
            ('planilla.termination',        [('employee_id', '=', emp.id)], 'liquidaciones'),
        ]
        for model, domain, label in related:
            try:
                recs = self.env[model].search(domain)
                if recs:
                    recs.unlink()
                    deleted.append(f'{len(recs)} {label}')
            except Exception:
                pass  # modelo puede no existir en esta instalacion

        # Eliminar el empleado
        emp_name = emp.name
        emp.unlink()
        deleted_txt = ', '.join(deleted) if deleted else 'sin registros adicionales'

        # Reabrir el wizard con el resultado
        self.write({'sample_exists': False})
        return {
            'type':    'ir.actions.act_window',
            'res_model': self._name,
            'res_id':  self.id,
            'view_mode': 'form',
            'view_id': self.env.ref('planilla_cr.view_import_data_wizard_form').id,
            'target': 'new',
            'context': {
                'default_delete_msg': f'OK Empleado "{emp_name}" eliminado correctamente. '
                                      f'Registros eliminados: {deleted_txt}.'
            },
        }
