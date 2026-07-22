from odoo import models, fields, api
import base64, io

# -- openpyxl -----------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation


class ImportTemplateWizard(models.TransientModel):
    """Genera y descarga el machote Excel para carga masiva de empleados."""
    _name        = 'planilla.import.template.wizard'
    _description = 'Machote de Importacion de Empleados'

    company_id = fields.Many2one(
        'res.company', string='Empresa', required=True,
        default=lambda self: self.env.company,
    )
    include_employees    = fields.Boolean('Empleados (datos principales)', default=True)
    include_loans        = fields.Boolean('Prestamos y Adelantos',         default=True)
    include_pension      = fields.Boolean('Pensiones Alimentarias',        default=True)
    include_benefits     = fields.Boolean('Beneficios / Deducciones Recurrentes', default=True)
    include_disabilities = fields.Boolean('Incapacidades',                 default=True)
    include_vacations    = fields.Boolean('Saldo de Vacaciones',           default=True)
    include_overtime     = fields.Boolean('  Horas Extras (historico)',   default=True)
    include_embargos     = fields.Boolean('  Embargos Judiciales',        default=True)
    include_bonos        = fields.Boolean('  Bonos y Beneficios',          default=True)
    include_cobros       = fields.Boolean('  Cobros a Empleados',             default=True)
    include_acumulados   = fields.Boolean('Acumulados Provisiones (aguinaldo/cesantia/vacaciones)', default=True)
    include_config       = fields.Boolean('Configuracion inicial (tramos renta, calendarizacion)', default=True)
    include_sample_data  = fields.Boolean(
        '  Incluir fila de prueba (EMPLEADO PRUEBA)',
        default=False,
        help='Agrega una fila naranja de prueba en todas las hojas con cedula 1-0000-0001. '
             'Active solo cuando quiera verificar que la importacion funciona correctamente. '
             'Luego use el boton "Eliminar Empleado de Prueba" para limpiar.'
    )

    # Cedula reservada para la fila de prueba -- misma en template y en import wizard
    _SAMPLE_CEDULA = '1-0000-0001'

    # -- Listas de valores para dropdowns Excel --------------------------------
    # Orden importa: cada key ocupa una columna en la hoja oculta _LISTAS
    _DV_LISTS = {
        # Identificacion
        'id_type':      ['Cedula Nacional', 'Residencia / DIMEX',
                         'Permiso de Trabajo', 'Pasaporte', 'Indocumentado'],
        # INS
        'ins_risk':     ['I - Oficinas', 'II - Comercio', 'III - Industria',
                         'IV - Construccion', 'V - Alto Riesgo'],
        'ins_workday':  ['Ordinaria', 'Extraordinaria', 'Mixta',
                         'Tiempo Parcial', 'Por Horas', 'Ocasional'],
        'ins_civil':    ['Soltero/a', 'Casado/a', 'Divorciado/a',
                         'Viudo/a', 'Union Libre', 'Separado/a'],
        'ins_nat':      ['Costarricense', 'Nicaraguense', 'Colombiano/a',
                         'Estadounidense', 'Hondureno/a', 'Salvadoreno/a',
                         'Guatemalteco/a', 'Panameno/a', 'Mexicano/a',
                         'Venezolano/a', 'Peruano/a', 'Ecuatoriano/a', 'Otra'],
        # Banco y cuenta
        'banco':        ['BNCR', 'BCR', 'BP', 'BAC', 'BCT', 'CATHAY', 'CMB',
                         'DAVIVIENDA', 'GENERAL', 'IMPROSA', 'LAFISE',
                         'PROMERICA', 'PRIVAL', 'SCOTIA', 'COOCIQUE',
                         'COOPENAE', 'MUTUAL_ALJ', 'Otro'],
        'account_type': ['Cuenta Corriente', 'Cuenta de Ahorros', 'SINPE Movil'],
        # Nomina
        'frequency':    ['Mensual', 'Quincenal', 'Semanal', 'Bimensual'],
        'calc_method':  ['Salario Fijo', 'Por Horas Trabajadas'],
        # Genero y si/no
        'gender':       ['Masculino', 'Femenino', 'Otro'],
        'si_no':        ['Si', 'No'],
        # Horarios -- sincronizado con default_data.xml
        # Prestamos
        'loan_type':    ['Prestamo de Empresa', 'Adelanto de Salario'],
        'loan_state':   ['Aprobado', 'En Curso', 'Borrador', 'Pagado', 'Anulado'],
        # Pension
        'pension_rel':  ['Hijo/a', 'Conyuge', 'Padre', 'Madre', 'Otro'],
        'pension_calc': ['Porcentaje del Salario', 'Monto Fijo'],
        # Beneficios
        'benefit_type': ['Beneficio / Ingreso', 'Deduccion / Descuento'],
        'amount_type':  ['Monto Fijo', 'Porcentaje'],
        # Incapacidades
        'disability':   ['Enfermedad Comun (CCSS)', 'Accidente de Trabajo (CCSS)',
                         'Riesgo Laboral (INS)', 'Maternidad / Paternidad', 'Otro'],
        # Horas extras
        'overtime_type':['Simple (1.5x)', 'Doble (2.0x)', 'Dia Feriado'],
        # Embargos
        'embargo_calc': ['Monto Fijo', 'Porcentaje del Neto Disponible'],
        # Bonos
        'bono_type':    ['Productividad / Rendimiento', 'Asistencia Perfecta',
                         'Antiguedad por Anos de Servicio', 'Subsidio de Transporte / Kilometraje',
                         'Subsidio de Alimentacion (en dinero)', 'Subsidio Educativo',
                         'Subsidio de Salud / Medico', 'Gastos de Representacion',
                         'Comision por Ventas', 'Incentivo / Premio Especial', 'Otro'],
        'bono_calc':    ['Monto Fijo', 'Porcentaje del Salario Base'],
        'si_no_recurrente': ['Si', 'No'],
        # Tipo de sangre
        'blood_type':   ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-'],
        # Ocupaciones INS -- COCR-2023 (INEC), basada en CIUO-08 OIT
        'ins_occupation': [
            '[1111] Miembros del poder legislativo y ejecutivo',
            '[1112] Personal directivo de la administracion publica',
            '[1113] Jefes de comunidades etnicas',
            '[1114] Dirigentes de organizaciones que presentan un interes especial',
            '[1120] Directores y gerentes generales',
            '[1211] Directores y gerentes de servicios financieros',
            '[1212] Directores y gerentes de recursos humanos',
            '[1213] Directores y gerentes de politicas y planificacion',
            '[1219] Directores y gerentes de administracion y servicios no clasificados bajo otros',
            '[1221] Directores y gerentes de venta y comercializacion',
            '[1222] Directores y gerentes de publicidad y relaciones publicas',
            '[1223] Directores y gerentes de investigacion y desarrollo',
            '[1311] Directores y gerentes de produccion agropecuaria y silvicultura',
            '[1312] Directores y gerentes de produccion acuicola, piscicola y de pesca',
            '[1321] Directores y gerentes de industrias manufactureras',
            '[1322] Directores y gerentes de explotaciones de mineria',
            '[1323] Directores y gerentes de empresas de construccion',
            '[1324] Directores y gerentes de empresas de abastecimiento, distribucion y afines',
            '[1330] Directores y gerentes de servicios de tecnologia de la informacion',
            '[1341] Directores y gerentes de servicios de cuidados infantiles',
            '[1342] Directores y gerentes de servicios de salud',
            '[1343] Directores y gerentes de servicios de atencion a personas adultas mayores',
            '[1344] Directores y gerentes de servicios de bienestar social',
            '[1345] Directores y gerentes de servicios de educacion',
            '[1346] Directores y gerentes de sucursales de bancos y servicios financieros',
            '[1349] Directores y gerentes de servicios profesionales no clasificados bajo otros',
            '[1411] Directores y gerentes de hoteles',
            '[1412] Directores y gerentes de restaurantes',
            '[1420] Gerentes de comercios al por mayor y al por menor',
            '[1431] Directores y gerentes de centros deportivos, de esparcimiento y culturales',
            '[1439] Directores y gerentes de servicios no clasificados bajo otros',
            '[2111] Fisicos y astronomos', '[2112] Meteorologos', '[2113] Quimicos',
            '[2114] Geologos y geofisicos', '[2120] Matematicos, actuarios y estadisticos',
            '[2131] Biologos, botanicos, zoologos y afines',
            '[2132] Agronomos, zootecnistas y afines',
            '[2133] Profesionales de la proteccion medioambiental',
            '[2141] Ingenieros industriales y de produccion', '[2142] Ingenieros civiles',
            '[2143] Ingenieros medioambientales',
            '[2144] Ingenieros mecanicos, navales y aeronauticos',
            '[2145] Ingenieros quimicos', '[2146] Ingenieros de minas, metalurgicos y afines',
            '[2149] Ingenieros no clasificados bajo otros epigrafes',
            '[2151] Ingenieros electricos', '[2152] Ingenieros electronicos',
            '[2153] Ingenieros en telecomunicaciones, audio y sonido',
            '[2161] Arquitectos', '[2162] Arquitectos paisajistas',
            '[2163] Disenadores industriales de productos y moda',
            '[2164] Urbanistas e ingenieros de transito', '[2165] Topografos',
            '[2166] Disenadores graficos y multimedia',
            '[2211] Medicos generales', '[2212] Medicos geriatras',
            '[2213] Medicos ginecologos y obstetras', '[2214] Medicos psiquiatras',
            '[2215] Medicos ortopedistas y traumatologos',
            '[2219] Especialistas medicos no clasificados bajo otros epigrafes',
            '[2220] Enfermeros profesionales y profesionales de parteria',
            '[2230] Profesionales de medicina tradicional y alternativa',
            '[2250] Veterinarios', '[2261] Dentistas',
            '[2262] Cirujanos orales y maxilofaciales', '[2271] Farmaceuticos',
            '[2272] Profesionales de la salud y la higiene laboral y ambiental',
            '[2273] Fisioterapeutas', '[2274] Nutricionistas',
            '[2275] Audiologos y terapeutas del lenguaje', '[2276] Optometristas',
            '[2279] Profesionales de la salud no clasificados bajo otros epigrafes',
            '[2310] Profesores de instituciones de educacion superior',
            '[2320] Profesores de formacion profesional',
            '[2330] Profesores de educacion secundaria',
            '[2341] Profesores de educacion primaria', '[2342] Profesores de educacion preescolar',
            '[2351] Especialistas en metodos pedagogicos',
            '[2352] Profesores de educacion especial', '[2353] Otros profesores de idiomas',
            '[2354] Otros profesores de musica', '[2355] Otros profesores de artes',
            '[2356] Instructores en tecnologia de la informacion',
            '[2359] Profesionales de la educacion no clasificados bajo otros epigrafes',
            '[2411] Contadores y auditores financieros',
            '[2412] Asesores financieros y en inversiones', '[2413] Analistas financieros',
            '[2421] Analistas de gestion y organizacion',
            '[2422] Profesionales en politicas sociales y de administracion',
            '[2423] Profesionales de gestion de talento humano',
            '[2424] Profesionales en formacion, desarrollo de personal y evaluacion',
            '[2431] Profesionales de la publicidad y la comercializacion',
            '[2432] Profesionales de relaciones publicas',
            '[2433] Profesionales de ventas tecnicas y medicas',
            '[2434] Profesionales de ventas de tecnologia de la informacion',
            '[2511] Analistas de sistemas', '[2512] Desarrolladores de software',
            '[2513] Desarrolladores web y multimedia', '[2514] Programadores de aplicaciones',
            '[2519] Desarrolladores y analistas de software no clasificados bajo otros',
            '[2521] Disenadores y administradores de bases de datos',
            '[2522] Administradores de sistemas', '[2523] Profesionales en redes de computadores',
            '[2529] Profesionales en bases de datos y redes no clasificados bajo otros',
            '[2611] Abogados', '[2612] Jueces',
            '[2619] Profesionales en derecho no clasificados bajo otros epigrafes',
            '[2621] Archivistas, curadores de arte y restauradores',
            '[2622] Bibliotecologos, documentalistas y afines',
            '[2631] Economistas', '[2632] Sociologos, antropologos y afines',
            '[2633] Filosofos, historiadores y especialistas en ciencias politicas',
            '[2634] Psicologos', '[2635] Profesionales del trabajo social',
            '[2636] Profesionales religiosos',
            '[2639] Profesionales en ciencias sociales no clasificados bajo otros',
            '[2641] Autores literarios y otros escritores',
            '[2642] Periodistas, editores y redactores',
            '[2643] Traductores, interpretes, linguistas y filologos',
            '[2651] Escultores, pintores artisticos y afines',
            '[2652] Musicos, cantantes y compositores',
            '[2653] Coreografos, directores de danza y bailarines profesionales',
            '[2654] Directores y productores de cine, de teatro y afines', '[2655] Actores',
            '[3111] Tecnicos en ciencias fisicas y quimicas', '[3112] Tecnicos en ingenieria civil',
            '[3113] Electrotecnicos', '[3114] Tecnicos en electronica',
            '[3115] Tecnicos en ingenieria mecanica', '[3116] Tecnicos en quimica industrial',
            '[3117] Tecnicos en ingenieria de minas y metalurgia',
            '[3118] Delineantes y dibujantes tecnicos',
            '[3119] Otros tecnicos en ciencias fisicas, quimica e ingenieria no clasificados',
            '[3121] Supervisores en ingenieria de minas',
            '[3122] Supervisores en industrias manufactureras',
            '[3123] Supervisores de la construccion',
            '[3131] Operadores de plantas de generacion y distribucion de energia',
            '[3132] Operadores de incineradores y plantas de tratamiento de agua',
            '[3133] Controladores de instalaciones de procesamiento de productos quimicos',
            '[3134] Operadores de instalaciones de refinacion de petroleo y gas natural',
            '[3135] Controladores de procesos de produccion de metales',
            '[3139] Tecnicos en control de procesos no clasificados bajo otros',
            '[3141] Tecnicos en ciencias biologicas',
            '[3142] Tecnicos agropecuarios', '[3143] Tecnicos forestales',
            '[3151] Maquinistas en navegacion maritima',
            '[3152] Capitanes y oficiales de cubierta',
            '[3153] Pilotos de aviacion y afines', '[3154] Controladores de trafico aereo',
            '[3155] Tecnicos en seguridad aeronautica',
            '[3211] Tecnicos en aparatos de diagnostico y tratamiento medico',
            '[3212] Tecnicos de laboratorios medicos',
            '[3213] Tecnicos y asistentes en farmacia',
            '[3214] Tecnicos de protesis medicas y dentales',
            '[3220] Profesionales de nivel medio de enfermeria',
            '[3230] Profesionales de nivel medio de medicina tradicional y alternativa',
            '[3240] Tecnicos y asistentes veterinarios',
            '[3250] Tecnico en emergencias medicas',
            '[3261] Auxiliares y tecnicos de odontologia',
            '[3262] Tecnicos en documentacion sanitaria',
            '[3263] Trabajadores comunitarios de la salud',
            '[3264] Tecnicos en optometria y opticos',
            '[3265] Tecnicos y asistentes fisioterapeutas',
            '[3266] Practicantes y asistentes medicos',
            '[3267] Inspectores de la salud laboral y medioambiental',
            '[3268] Auxiliar de ambulancias en emergencias medicas',
            '[3269] Tecnicos de las ciencias de la salud no clasificados bajo otros',
            '[3311] Agentes de bolsa, cambio y otros servicios financieros',
            '[3312] Oficiales de prestamos y creditos',
            '[3313] Tecnicos y auxiliares de contabilidad',
            '[3314] Profesionales de nivel medio de servicios estadisticos y matematicos',
            '[3315] Tasadores', '[3316] Tecnicos y asistentes en administracion y economia',
            '[3321] Agentes de seguros', '[3322] Representantes comerciales',
            '[3323] Agentes de proveeduria', '[3324] Agentes de compras y consignatarios',
            '[3331] Declarantes o gestores de aduana',
            '[3332] Organizadores de conferencias y eventos',
            '[3333] Agentes de empleo y contratistas de mano de obra',
            '[3334] Agentes inmobiliarios',
            '[3339] Otros agentes comerciales y corredores no clasificados bajo otros',
            '[3341] Supervisores de oficina', '[3342] Secretarios juridicos',
            '[3343] Secretarios administrativos y ejecutivos', '[3344] Secretarios medicos',
            '[3351] Inspectores de aduanas y fronteras',
            '[3352] Agentes de administracion tributaria',
            '[3353] Agentes de servicios de seguridad social',
            '[3354] Funcionarios de servicios de expedicion de licencias y permisos',
            '[3355] Inspectores de policia y detectives',
            '[3411] Profesionales de nivel medio del derecho y servicios legales',
            '[3412] Tecnicos y asistentes en trabajo social',
            '[3421] Atletas y deportistas',
            '[3422] Entrenadores, instructores y arbitros de actividades deportivas',
            '[3423] Instructores de educacion fisica y actividades recreativas',
            '[3431] Fotografos', '[3432] Disenadores y decoradores de interior',
            '[3511] Tecnicos en operaciones de tecnologia de la informacion',
            '[3512] Tecnicos en asistencia al usuario de tecnologia de la informacion',
            '[3513] Tecnicos en redes y sistemas de computadores', '[3514] Tecnicos de la web',
            '[3521] Tecnicos de radiodifusion y grabacion audiovisual',
            '[3522] Tecnicos de ingenieria de las telecomunicaciones',
            '[4110] Oficinistas generales', '[4120] Secretarios generales',
            '[4131] Operadores de maquinas de procesamiento de texto y mecanografos',
            '[4132] Digitadores de datos', '[4211] Cajeros de bancos y afines',
            '[4221] Recepcionistas',
            '[4222] Empleados de atencion y asesoramiento de llamadas',
            '[4229] Empleados de servicios de informacion al cliente no clasificados bajo otros',
            '[4311] Empleados de contabilidad y calculo de costos',
            '[4312] Empleados de servicios estadisticos, financieros y de seguros',
            '[4313] Empleados encargados de las nominas',
            '[4321] Empleados de control de abastecimientos e inventario',
            '[4322] Empleados de servicios de apoyo a la produccion',
            '[4323] Empleados de servicio de transporte',
            '[5111] Auxiliares de servicio abordo', '[5120] Cocineros', '[5131] Saloneros',
            '[5132] Bartenders',
            '[5141] Especialistas en tratamientos del cabello',
            '[5142] Especialistas en tratamientos de belleza estetica y afines',
            '[5151] Supervisores limpieza en oficinas, hoteles y otros establecimientos',
            '[5153] Encargados de mantenimiento de edificios',
            '[5165] Instructores de manejo', '[5169] Otros trabajadores de servicios personales',
            '[5211] Vendedores de quioscos y de puestos de mercado',
            '[5221] Propietarios y comerciantes encargados de pequenas tiendas',
            '[5222] Supervisores de tiendas y almacenes',
            '[5223] Asistentes de ventas de tiendas y almacenes',
            '[5230] Cajeros y expendedores de boletos y tiquetes',
            '[5243] Vendedores puerta a puerta', '[5244] Vendedores por telefono',
            '[5246] Vendedores de comidas al mostrador',
            '[5249] Vendedores no clasificados bajo otros epigrafes',
            '[5311] Cuidadores de ninos', '[5312] Ayudantes de maestros',
            '[5321] Trabajadores de los cuidados personales en instituciones',
            '[5322] Trabajadores de los cuidados personales a domicilio',
            '[5411] Bomberos', '[5412] Policias e inspectores de transito',
            '[5414] Guardas de proteccion en establecimientos',
            '[5415] Vigilante de casas particulares',
            '[5419] Otros trabajadores que prestan servicios de proteccion y vigilancia',
            '[6111] Agricultores y trabajadores calificados de cultivos',
            '[6121] Criadores de ganado', '[6122] Avicultores y trabajadores de avicultura',
            '[6130] Productores y trabajadores de explotaciones agropecuarias mixtas',
            '[6210] Trabajadores forestales calificados y afines',
            '[7111] Albaniles',
            '[7113] Operarios en cemento armado, encofradores y afines',
            '[7114] Carpinteros de armar y de obra blanca',
            '[7121] Techadores', '[7122] Revestidores e instaladores de pisos',
            '[7126] Fontaneros e instaladores de tuberias',
            '[7127] Mecanicos de instalaciones de refrigeracion y aire acondicionado',
            '[7131] Pintores y empapeladores',
            '[7211] Moldeadores de metal', '[7212] Soldadores y oxicortadores',
            '[7214] Montadores de estructuras metalicas',
            '[7221] Herreros y forjadores', '[7222] Herramentistas y afines',
            '[7223] Reguladores y operadores de maquinas herramientas',
            '[7231] Mecanicos y reparadores de vehiculos de motor',
            '[7232] Mecanicos y reparadores de motores de avion',
            '[7233] Mecanicos y reparadores de maquinas agricolas e industriales',
            '[7311] Mecanicos y reparadores de instrumentos de precision',
            '[7313] Joyeros, orfebres y plateros',
            '[7411] Electricistas de obras y afines',
            '[7412] Mecanicos y ajustadores electricistas',
            '[7413] Instaladores y reparadores de lineas electricas',
            '[7421] Mecanicos y reparadores en electronica',
            '[7422] Instaladores y reparadores en tecnologia de la informacion',
            '[7511] Carniceros, pescadores y afines',
            '[7512] Panaderos, pasteleros, golosineros y confiteros',
            '[8111] Mineros y operadores de instalaciones mineras',
            '[8211] Ensambladores de maquinaria mecanica',
            '[8212] Ensambladores de equipos electricos y electronicos',
            '[8322] Conductores de automoviles, taxis y camionetas',
            '[8331] Conductores de autobuses y tranvias',
            '[8332] Conductores de camiones pesados',
            '[8341] Operadores de maquinaria agricola y forestal movil',
            '[8342] Operadores de maquinas de movimiento de tierras y afines',
            '[8343] Operadores de gruas, aparatos elevadores y afines',
            '[9111] Limpiadores y asistentes domesticos',
            '[9112] Limpiadores y asistentes de oficinas, hoteles y otros establecimientos',
            '[9211] Peones de explotaciones agricolas',
            '[9311] Peones de minas y canteras',
            '[9312] Peones de obras publicas y mantenimiento',
            '[9313] Peones de la construccion de edificios',
            '[9321] Empacadores manuales',
            '[9329] Peones de la industria manufacturera no clasificados bajo otros',
            '[9333] Peones de carga', '[9334] Reponedores de estanterias',
            '[9411] Cocineros de comidas rapidas', '[9412] Ayudantes de cocina',
            '[9611] Recolectores de basura y material reciclable',
            '[9621] Mensajeros, mandaderos, maleteros y repartidores',
            '[9629] Ocupaciones elementales no clasificadas bajo otros epigrafes',
        ],
    }

    # -- paleta ----------------------------------------------------------------
    _C = {
        'dark':     '1F3864',
        'med':      '2E75B6',
        'light':    'D6E4F0',
        'req':      'FFF2CC',
        'opt':      'FFFFFF',
        'example':  'E2EFDA',
        'border':   'BDD7EE',
        'white':    'FFFFFF',
        'red_hdr':  'C00000',
    }

    # -- catalogos para dropdowns ----------------------------------------------
    # Orden de columnas en la hoja oculta _LISTAS (A, B, C, ...)
    # Clave -> (col_idx_0based, [valores])
    _LISTAS = {
        'id_type':          (0,  ['01','02','03','04','05']),
        'si_no':            (1,  ['si','no']),
        'ins_risk':         (2,  ['I','II','III','IV','V']),
        'ins_workday':      (3,  ['01','02','03','04','05','06']),
        'banco':            (4,  ['BNCR','BCR','BP','BAC','BCT','CATHAY','CMB',
                                  'DAVIVIENDA','GENERAL','IMPROSA','LAFISE',
                                  'PROMERICA','PRIVAL','SCOTIA','COOCIQUE',
                                  'COOPENAE','MUTUAL_ALJ','OTRO']),
        'account_type':     (5,  ['corriente','ahorros','sinpe']),
        'frequency':        (6,  ['monthly','biweekly','weekly','bimonthly']),
        'calc_method':      (7,  ['fixed','attendance']),
        'ins_nationality':  (8,  ['CR','NI','CO','US','HN','SV','GT','PA',
                                  'MX','VE','PE','EC','OT']),
        'ins_civil':        (9,  ['01','02','03','04','05','06']),
        'gender':           (10, ['masculino','femenino','otro']),
        'loan_type':        (11, ['loan','advance']),
        'loan_state':       (12, ['approved','active','draft','paid','cancelled']),
        'pension_relacion': (13, ['hijo','conyuge','padre','madre','otro']),
        'pension_calc':     (14, ['porcentaje','monto_fijo']),
        'benefit_type':     (15, ['beneficio','deduccion']),
        'amount_type':      (16, ['fijo','porcentaje']),
        'disability_type':  (17, ['ccss','ccss_accident','ins','maternity','other']),
        'overtime_type':    (18, ['simple','double','holiday']),
    }

    # -- helpers ---------------------------------------------------------------
    @staticmethod
    def _fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    @staticmethod
    def _font(bold=False, color='000000', size=10, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic, name='Arial')

    @staticmethod
    def _border():
        s = Side(style='thin', color='BDD7EE')
        return Border(left=s, right=s, top=s, bottom=s)

    @staticmethod
    def _center():
        return Alignment(horizontal='center', vertical='center', wrap_text=True)

    @staticmethod
    def _left():
        return Alignment(horizontal='left', vertical='center', wrap_text=True)

    def _hdr(self, cell, text, bg=None, txt='FFFFFF', bold=True, size=10):
        C = self._C
        cell.value     = text
        cell.font      = self._font(bold=bold, color=txt, size=size)
        cell.fill      = self._fill(bg or C['med'])
        cell.border    = self._border()
        cell.alignment = self._center()

    def _col_hdr(self, cell, text, required, desc=''):
        C = self._C
        bg  = C['req'] if required else C['light']
        txt = '7B3F00' if required else C['dark']
        cell.value     = text
        cell.font      = self._font(bold=True, color=txt, size=9)
        cell.fill      = self._fill(bg)
        cell.border    = self._border()
        cell.alignment = self._center()
        if desc:
            c = Comment(desc, 'Planilla CR')
            c.width, c.height = 220, 55
            cell.comment = c

    def _example(self, cell, value):
        cell.value     = value
        cell.fill      = self._fill(self._C['example'])
        cell.border    = self._border()
        cell.alignment = self._left()
        cell.font      = self._font(italic=True, size=9, color='375623')

    def _data(self, cell, required=True):
        bg = self._C['req'] if required else self._C['opt']
        cell.fill      = self._fill(bg)
        cell.border    = self._border()
        cell.alignment = self._left()
        cell.font      = self._font(size=10)

    @staticmethod
    def _w(ws, col_idx, width):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # -- titulo de hoja --------------------------------------------------------
    def _sheet_title(self, ws, text, ncols, bg=None):
        C = self._C
        col_letter = get_column_letter(ncols)
        ws.merge_cells(f'A1:{col_letter}1')
        c = ws['A1']
        c.value     = text
        c.font      = self._font(bold=True, color=C['white'], size=12)
        c.fill      = self._fill(bg or C['dark'])
        c.alignment = self._center()
        ws.row_dimensions[1].height = 28

    # -- filas de datos (vacias + ejemplo + prueba) -----------------------------
    def _build_rows(self, ws, cols, data_rows=80, header_row=2, example_row=3,
                    sample_values=None):
        """
        cols = [(nombre, required, width, ejemplo, desc), ...]
        sample_values: lista de valores para la fila de prueba (naranja).
                       Si es None, no se dibuja fila de prueba.
        """
        # Encabezados
        for ci, (nombre, req, w, _, desc) in enumerate(cols, 1):
            self._col_hdr(ws.cell(header_row, ci), nombre, req, desc)
            self._w(ws, ci, w)
        ws.row_dimensions[header_row].height = 45

        # Fila de ejemplo (verde)
        for ci, (_, _, _, ej, _) in enumerate(cols, 1):
            self._example(ws.cell(example_row, ci), ej)
        ws.row_dimensions[example_row].height = 16

        # Fila de prueba (naranja) -- opcional, justo debajo del ejemplo
        data_start = example_row + 1
        if sample_values:
            sample_row = example_row + 1
            data_start = sample_row + 1
            for ci, val in enumerate(sample_values, 1):
                self._sample(ws.cell(sample_row, ci), val)
            ws.row_dimensions[sample_row].height = 16

        # Filas vacias
        for r in range(data_start, data_start + data_rows):
            for ci, (_, req, _, _, _) in enumerate(cols, 1):
                self._data(ws.cell(r, ci), req)
            ws.row_dimensions[r].height = 16

        ws.freeze_panes = ws.cell(example_row, 1)
        ws.sheet_view.showGridLines = False

    def _sample(self, cell, value):
        """Estilo para la fila de prueba: fondo naranja, texto oscuro, italica."""
        cell.value     = value
        cell.fill      = self._fill('F4B942')
        cell.border    = self._border()
        cell.alignment = self._left()
        cell.font      = self._font(italic=True, bold=True, size=9, color='7B2D00')

    def _build_listas_sheet(self, wb):
        """Crea hoja de listas para DataValidation.
        IMPORTANTE: la hoja se deja VISIBLE (no oculta) porque Excel 2016/2019
        y algunas versiones de Excel Online no muestran el dropdown cuando la
        hoja fuente esta oculta. Se protege y se estiliza como hoja de sistema
        para que el usuario no la modifique accidentalmente.
        """
        ws = wb.create_sheet(' LISTAS')
        C = self._C

        # Encabezado de advertencia fila 1
        ws.merge_cells('A1:Z1')
        c = ws['A1']
        c.value  = 'WARN  HOJA DE SISTEMA -- No modificar. Contiene las listas de los desplegables.'
        c.font   = self._font(bold=True, color='FFFFFF', size=9)
        c.fill   = self._fill('7F7F7F')
        c.alignment = self._center()
        ws.row_dimensions[1].height = 18

        # Escribir cada lista empezando en fila 2 con header de columna
        for col_idx, (key, vals) in enumerate(self._DV_LISTS.items(), 1):
            # Header de la columna
            hdr = ws.cell(2, col_idx)
            hdr.value = key
            hdr.font  = self._font(bold=True, color='FFFFFF', size=8)
            hdr.fill  = self._fill('404040')
            hdr.alignment = self._center()
            # Valores desde fila 3
            for row_idx, val in enumerate(vals, 3):
                cell = ws.cell(row_idx, col_idx, value=val)
                cell.font = Font(name='Arial', size=8, color='333333')
                cell.fill = PatternFill('solid', fgColor='F2F2F2')

        # Ajustar anchos minimos
        ws.column_dimensions['A'].width = 20
        # Columna de ocupaciones -- mas ancha
        keys = list(self._DV_LISTS.keys())
        if 'ins_occupation' in keys:
            occ_col = get_column_letter(keys.index('ins_occupation') + 1)
            ws.column_dimensions[occ_col].width = 70

        # Proteger la hoja para que no se edite accidentalmente
        ws.protection.sheet     = True
        ws.protection.password  = 'planilla_cr_sys'
        ws.protection.enable()

        ws.sheet_properties.tabColor = '808080'

        return ws, len(self._DV_LISTS)  # retorna nro de columnas estaticas

    def _build_dynamic_lists(self, wb, company_id, static_cols):
        """Agrega listas dinamicas (desde BD) a la hoja  LISTAS.
        Se llama desde action_generate DESPUES de _build_listas_sheet.
        Retorna dict {clave: (col_letter, first_row, last_row)} para _dv_dynamic.
        """
        ws = wb[' LISTAS']
        ws.protection.sheet = False  # desproteger temporalmente para escribir

        next_col = static_cols + 1  # columna donde empiezan las listas dinamicas

        HDR_FONT  = self._font(bold=True, color='FFFFFF', size=8)
        HDR_FILL  = self._fill('1F4E79')
        GREY_FILL = self._fill('999999')
        DATA_FONT = Font(name='Arial', size=8, color='333333')
        DATA_FILL = PatternFill('solid', fgColor='EBF3FB')
        GREY_FONT = Font(name='Arial', size=8, color='999999', italic=True)

        co = company_id
        dyn = {}  # key -> (col_letter, first_row, last_row)

        def _write_list(key, values, width=35):
            """Escribe la lista en la columna actual.
            SIEMPRE incrementa next_col aunque values este vacio,
            para que las columnas siguientes no se desplacen.
            """
            nonlocal next_col
            col_letter = get_column_letter(next_col)
            hdr = ws.cell(2, next_col)
            hdr.value = key
            hdr.font  = HDR_FONT if values else GREY_FONT
            hdr.fill  = HDR_FILL if values else GREY_FILL
            hdr.alignment = self._center()
            if values:
                first_r = 3
                for i, val in enumerate(values, first_r):
                    c = ws.cell(i, next_col, value=val)
                    c.font = DATA_FONT
                    c.fill = DATA_FILL
                last_r = first_r + len(values) - 1
                ws.column_dimensions[col_letter].width = width
                dyn[key] = (col_letter, first_r, last_r)
            else:
                ws.cell(3, next_col, value='(sin registros -- creelos en Odoo primero)')
                ws.cell(3, next_col).font = GREY_FONT
                ws.column_dimensions[col_letter].width = 28
            next_col += 1  # siempre avanzar

        # Usamos sudo() para bypassear las reglas multi-empresa del ORM
        # (el usuario que arma el import puede no tener acceso directo a
        # planilla.branch, planilla.calendar, etc. de su propia compañía).
        # BUG FIX: el filtro de compañía se aplica AQUI DENTRO, dentro del
        # helper -- antes se dejaba como responsabilidad de cada llamada,
        # pero ninguna de las 9 llamadas lo pasaba, asi que la plantilla
        # traia sucursales/departamentos/puestos/calendarios de TODAS
        # las compañías de la base de datos.
        def _search(model, domain=None, order='name'):
            dom = list(domain or [])
            if 'company_id' in self.env[model]._fields:
                dom += ['|',
                    ('company_id', '=', self.company_id.id),
                    ('company_id', '=', False)]
            return self.env[model].sudo().with_context(active_test=False).search(
                dom, order=order)

        # -- Tipos de horario ---------------------------------------------
        schedules = _search('planilla.schedule.type')
        _write_list('schedule', [s.name for s in schedules], width=40)

        # -- Calendarizaciones de planilla ---------------------------------
        # Sin filtro de empresa: sudo() ya bypasea ir.rules. En un sistema
        # de una sola empresa todos los registros son del cliente.
        cals = _search('planilla.calendar')
        _write_list('calendar', [c.name for c in cals], width=28)

        # -- Tipos de empleado ---------------------------------------------
        etypes = _search('planilla.employee.type')
        _write_list('employee_type', [e.name for e in etypes], width=28)

        # -- Estados de empleado -------------------------------------------
        estatuses = _search('planilla.employee.status')
        _write_list('employee_status', [e.name for e in estatuses], width=24)

        # -- Sucursales ----------------------------------------------------
        branches = _search('planilla.branch')
        _write_list('branch', [b.name for b in branches], width=28)

        # -- Departamentos -------------------------------------------------
        depts = _search('hr.department', [('parent_id', '=', False)])
        _write_list('department', [d.name for d in depts], width=30)

        # -- Sub-departamentos ---------------------------------------------
        subdepts = _search('hr.department', [('parent_id', '!=', False)])
        _write_list('subdepartment', [d.name for d in subdepts], width=30)

        # -- Puestos / Cargos ----------------------------------------------
        jobs = _search('hr.job')
        _write_list('job', [j.name for j in jobs], width=28)

        # -- Paises --------------------------------------------------------
        countries = self.env['res.country'].search([], order='name')
        _write_list('country', [c.name for c in countries], width=28)

        # Re-proteger
        ws.protection.sheet    = True
        ws.protection.password = 'planilla_cr_sys'
        ws.protection.enable()

        return dyn  # dict {key: (col_letter, first_row, last_row)}


    def _dv_dynamic(self, ws, col_idx, dyn_key, first_data_row,
                    dyn_lists, last_data_row=500, title='Opciones'):
        """Aplica dropdown usando una lista dinamica (de BD) en  LISTAS.
        dyn_lists: dict retornado por _build_dynamic_lists().
        """
        dyn = dyn_lists or {}
        if dyn_key not in dyn:
            return  # catalogo vacio en BD -- no agrega dropdown
        col_letter_src, first_r, last_r = dyn[dyn_key]
        formula    = f"' LISTAS'!${col_letter_src}${first_r}:${col_letter_src}${last_r}"
        col_letter = get_column_letter(col_idx)
        sqref      = f'{col_letter}{first_data_row}:{col_letter}{last_data_row}'
        dv = DataValidation(
            type='list',
            formula1=formula,
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorStyle='warning',
            errorTitle='Valor no reconocido',
            error='Seleccione un valor de la lista o creelo en Odoo primero.',
            showInputMessage=True,
            promptTitle=title,
            prompt='Seleccione de la lista (opciones cargadas desde Odoo)',
        )
        ws.add_data_validation(dv)
        dv.sqref = sqref

    def _dv(self, ws, col_idx, list_key, first_data_row, last_data_row=500,
            title='Opciones'):
        """Helper rapido que busca la lista en _DV_LISTS y aplica el dropdown.
        La hoja  LISTAS tiene: fila 1 = advertencia, fila 2 = headers,
        fila 3 en adelante = valores. Por eso el rango empieza en $3.
        """
        vals = self._DV_LISTS.get(list_key, [])
        if not vals:
            return
        keys = list(self._DV_LISTS.keys())
        listas_col = get_column_letter(keys.index(list_key) + 1)
        first_r    = 3                      # fila 3: primera fila de datos
        last_r     = 3 + len(vals) - 1      # ultima fila de datos
        formula    = f"' LISTAS'!${listas_col}${first_r}:${listas_col}${last_r}"
        col_letter = get_column_letter(col_idx)
        sqref      = f'{col_letter}{first_data_row}:{col_letter}{last_data_row}'

        dv = DataValidation(
            type='list',
            formula1=formula,
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorStyle='warning',
            errorTitle='Valor no reconocido',
            error='El valor ingresado no esta en el catalogo. Revise la hoja  LISTAS.',
            showInputMessage=True,
            promptTitle=title,
            prompt=f'Seleccione: {", ".join(str(v) for v in vals[:4])}{"..." if len(vals) > 4 else ""}',
        )
        ws.add_data_validation(dv)
        dv.sqref = sqref

    # ==========================================================================
    # HOJA INSTRUCCIONES
    # ==========================================================================
    def _build_instructions(self, wb):
        C = self._C
        ws = wb.active
        ws.title = ' INSTRUCCIONES'
        ws.sheet_view.showGridLines = False

        ws.merge_cells('A1:H1')
        c = ws['A1']
        c.value     = (f'MACHOTE DE IMPORTACION -- SISTEMA PLANILLA v5.28.58  '
                       f'|  {self.company_id.name}  |  Legislacion CR 2026')
        c.font      = self._font(bold=True, color=C['white'], size=13)
        c.fill      = self._fill(C['dark'])
        c.alignment = self._center()
        ws.row_dimensions[1].height = 34

        ws.merge_cells('A2:H2')
        c = ws['A2']
        c.value     = 'Carga masiva de empleados -- complete las hojas y entregue al implementador'
        c.font      = self._font(italic=True, color=C['white'], size=10)
        c.fill      = self._fill(C['med'])
        c.alignment = self._center()
        ws.row_dimensions[2].height = 18

        lines = [
            ('', ''),
            ('QUE ES ESTE ARCHIVO', ''),
            ('', 'Permite cargar todos los empleados y sus datos al modulo Planilla CR de Odoo '
                 'de una sola vez, evitando la digitacion manual uno a uno.'),
            ('', ''),
            ('HOJAS INCLUIDAS', ''),
            ('', '  EMPLEADOS            -> Datos principales (obligatorio completar)'),
            ('', '  PRESTAMOS            -> Prestamos y adelantos activos del empleado'),
            ('', '  PENSION_ALIMENTARIA   -> Ordenes judiciales de pension alimentaria'),
            ('', '  OTROS DESCUENTOS      -> Cuota sindical, cooperativa, ahorro voluntario, seguro medico (no embargos ni bonos formales)'),
            ('', '  INCAPACIDADES        -> Incapacidades activas al momento de la carga'),
            ('', '  VACACIONES           -> Saldo de vacaciones acumulado'),
            ('', '  HORAS EXTRAS         -> Horas extras historicas'),
            ('', '  EMBARGOS             -> Embargos judiciales (Art. 172 CT -- max. 25% neto)'),
            ('', '  BONOS                -> Bonos e incentivos (productividad, transporte, etc.)'),
            ('', '  CATALOGOS            -> Valores validos para campos de lista (NO editar)'),
            ('', ''),
            ('INSTRUCCIONES', ''),
            ('', '1. Complete la hoja EMPLEADOS -- un empleado por fila.'),
            ('', '2. Use la cedula como llave: debe coincidir exactamente en todas las hojas.'),
            ('', '3. Para prestamos, pensiones o beneficios multiples: agregue una fila por cada uno.'),
            ('', '4. Los campos de seleccion tienen menu desplegable -- haga clic en la celda y elija de la lista.'),
            ('', '5. Fechas en formato DD/MM/AAAA  (ejemplo: 15/03/2020).'),
            ('', '6. Montos en colones (CRC), sin simbolo ni comas  (ejemplo: 750000).'),
            ('', '7. La fila de PRUEBA (fondo naranja, cedula 1-0000-0001) sirve para verificar que la importacion funciona. Elimine ese empleado luego.'),
            ('', '8. NO modifique los encabezados ni el nombre de las hojas.'),
            ('', ''),
            ('CODIGO DE COLORES', ''),
        ]

        for i, (label, text) in enumerate(lines, start=3):
            ws.row_dimensions[i].height = 18
            if label:
                ws.merge_cells(f'A{i}:H{i}')
                c = ws.cell(i, 1, value=label)
                c.font      = self._font(bold=True, color=C['dark'], size=10)
                c.fill      = self._fill(C['light'])
                c.alignment = self._left()
            else:
                ws.merge_cells(f'B{i}:H{i}')
                ws.cell(i, 2, value=text).font = self._font(size=10)

        # leyenda colores
        last = ws.max_row + 1
        leyenda = [
            (C['req'],     ' Fondo AMARILLO -> Campo OBLIGATORIO'),
            (C['opt'],     ' Fondo BLANCO   -> Campo OPCIONAL'),
            (C['example'], ' Fondo VERDE    -> Fila de EJEMPLO (solo referencia, no se importa)'),
            ('F4B942',     ' Fondo NARANJA  -> Fila de PRUEBA (cedula 1-0000-0001) -- importar para verificar, luego eliminar'),
        ]
        for offset, (color, texto) in enumerate(leyenda):
            r = last + offset
            ws.cell(r, 1).fill   = self._fill(color)
            ws.cell(r, 1).border = self._border()
            ws.merge_cells(f'B{r}:H{r}')
            c = ws.cell(r, 2, value=texto)
            c.font = self._font(size=10)
            ws.row_dimensions[r].height = 18

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 90

    # ==========================================================================
    # HOJA EMPLEADOS
    # ==========================================================================
    def _build_employees(self, wb, sample=False, dyn_lists=None):
        ws = wb.create_sheet(' EMPLEADOS')
        self._sheet_title(ws, 'DATOS DE EMPLEADOS -- Un empleado por fila', 43)

        # Secciones (fila 2) -- 43 columnas totales
        secciones = [
            (1,  6,  'IDENTIFICACION'),
            (7,  14, 'DATOS LABORALES'),
            (15, 23, 'DATOS INS'),
            (24, 29, 'CCSS Y BANCO'),
            (30, 34, 'CONFIGURACION NOMINA'),
            (35, 43, 'DATOS PERSONALES Y MEDICOS'),
        ]
        for cs, ce, titulo in secciones:
            ws.merge_cells(start_row=2, start_column=cs,
                           end_row=2,   end_column=ce)
            self._hdr(ws.cell(2, cs), titulo)
        ws.row_dimensions[2].height = 18

        cols = [
            # -- Identificacion (cols 1-6) ---------------------------------
            ('Nombre Completo',           True,  28, 'Juan Perez Rodriguez',
             'Nombre completo del empleado'),
            ('Cedula / Identificacion',   True,  18, '1-2345-6789',
             'Cedula, DIMEX o pasaporte -- llave entre hojas'),
            ('Tipo de Identificacion',    True,  18, 'Cedula Nacional',
             'Seleccione del desplegable'),
            ('Fecha de Ingreso',          True,  14, '01/03/2020',
             'Formato DD/MM/AAAA'),
            ('Fecha de Salida',           False, 14, '',
             'Solo si ya no trabaja en la empresa'),
            ('Correo Corporativo',        False, 28, 'juan.perez@empresa.com',
             'Email de trabajo'),
            # -- Datos laborales (cols 7-14) -------------------------------
            ('Departamento',              False, 26, '',
             'Seleccione del desplegable (cargado desde Odoo)'),
            ('Sub Departamento',          False, 26, '',
             'Seleccione del desplegable (cargado desde Odoo)'),
            ('Sucursal',                  False, 22, '',
             'Seleccione del desplegable (cargado desde Odoo)'),
            ('Puesto / Cargo',            False, 24, '',
             'Seleccione del desplegable o escriba el nombre exacto'),
            ('Tipo de Empleado',          True,  22, 'Empleado Indefinido',
             'Seleccione del desplegable (cargado desde Odoo)'),
            ('Estado del Empleado',       True,  18, '',
             'Seleccione del desplegable (cargado desde Odoo)'),
            ('Tipo de Horario',           True,  32, '',
             'Seleccione del desplegable (cargado desde Odoo -- tipos de su empresa)'),
            ('Calendarizacion de Planilla', True, 26, '',
             'Seleccione del desplegable (cargado desde Odoo) -- Ej: Mensual, Quincenal'),
            # -- Datos INS (cols 15-23) ------------------------------------
            ('Incluir en INS',            True,  12, 'Si',
             'Si / No'),
            ('Numero de Poliza INS',      False, 18, 'POL-12345',
             'Numero de poliza del INS'),
            ('Nombre INS',                False, 18, 'Juan',
             'Nombre como aparece en el sistema INS'),
            ('Primer Apellido INS',       False, 16, 'Perez',    ''),
            ('Segundo Apellido INS',      False, 16, 'Rodriguez',''),
            ('Clase de Riesgo INS',       True,  22, 'I - Oficinas',
             'Seleccione del desplegable'),
            ('Jornada INS',               True,  18, 'Ordinaria',
             'Seleccione del desplegable'),
            ('Ocupacion INS',             True,  50, '[4110] Oficinistas generales',
             'Seleccione del desplegable -- COCR-2023 (INEC)'),
            ('Tipo de Sangre',            False, 10, 'O+',
             'Seleccione del desplegable: A+, A-, B+, B-, AB+, AB-, O+, O-'),
            # -- CCSS y banco (cols 24-29) ---------------------------------
            ('Numero CCSS',               False, 16, '123456789',
             'Numero de asegurado CCSS'),
            ('Asegurado CCSS',            True,  14, 'Si',
             'Si / No'),
            ('Cuenta Bancaria / IBAN',    False, 30, 'CR65015200000000000000',
             'IBAN de 22 caracteres'),
            ('SINPE Movil',               False, 14, '88887777',
             'Telefono de 8 digitos registrado en SINPE Movil'),
            ('Banco',                     False, 20, 'BNCR',
             'Seleccione del desplegable'),
            ('Tipo de Cuenta Banco',      False, 16, 'Cuenta Corriente',
             'Seleccione del desplegable'),
            # -- Configuracion nomina (cols 30-34) -------------------------
            ('Salario Base (CRC)',          True,  18, '750000',
             'Salario mensual en colones, sin comas ni simbolo'),
            ('Fecha Vigencia Salarial',   False, 18, '01/01/2026',
             'Desde cuando aplica el salario (DD/MM/AAAA)'),
            ('Metodo de Calculo',         True,  18, 'Salario Fijo',
             'Seleccione del desplegable'),
            ('Nacionalidad INS',          False, 14, 'Costarricense',
             'Seleccione del desplegable'),
            ('Salario Variable',          True,  16, 'No',
             'Si / No -- Active "Si" si recibe comisiones o HE recurrentes (Art. 153 CT).'),
            # -- Datos personales y medicos (cols 36-44) -------------------
            ('Estado Civil INS',          False, 16, 'Soltero/a',
             'Seleccione del desplegable'),
            ('Genero',                    False, 14, 'Masculino',
             'Seleccione del desplegable'),
            ('Pais',                      False, 22, 'Costa Rica',
             'Seleccione del desplegable -- pais de residencia'),
            ('Numero de Dependientes',    False, 12, '0',
             'Hijos u otros dependientes'),
            ('Direccion',                 False, 30, 'San Jose, Escazu',
             'Direccion de habitacion'),
            ('Telefono Personal',         False, 14, '88887777',
             'Numero de telefono personal'),
            ('Correo Personal',           False, 26, 'juan@gmail.com',
             'Correo electronico personal (privado)'),
            ('Diagnostico / Notas Medicas', False, 40, '',
             'Condiciones, alergias, medicamentos u otras notas para el INS o emergencias.'),
            ('Observaciones',             False, 30, '',
             'Notas internas del empleado'),
        ]

        sv = None
        if sample:
            sv = [
                # -- Identificacion (cols 1-6)
                'Juan Perez Rodriguez',      # Nombre Completo
                self._SAMPLE_CEDULA,          # Cedula (1-2345-6789)
                'Cedula Nacional',            # Tipo de Identificacion
                '15/03/2021',                 # Fecha de Ingreso
                '',                           # Fecha de Salida (activo)
                'jperez@empresa.com',         # Correo Corporativo
                # -- Datos laborales (cols 7-14)
                'Administracion',             # Departamento
                'Contabilidad',               # Sub Departamento
                'Principal',                  # Sucursal
                'Asistente Administrativo',   # Puesto / Cargo
                'Empleado Indefinido',        # Tipo de Empleado
                'Activo',                     # Estado del Empleado
                'Jornada Completa (8 horas - Lun a Vie)',  # Tipo de Horario -- seleccione del desplegable
                'Quincenal',                  # Calendarizacion de Planilla
                # -- Datos INS (cols 15-23)
                'Si',                         # Incluir en INS
                'POL-2025-00123',             # Numero de Poliza INS
                'Juan',                       # Nombre INS
                'Perez',                      # Primer Apellido INS
                'Rodriguez',                  # Segundo Apellido INS
                'I - Oficinas',               # Clase de Riesgo INS
                'Ordinaria',                  # Jornada INS
                '[4313] Empleados encargados de las nominas',  # Ocupacion INS
                'O+',                         # Tipo de Sangre
                # -- CCSS y banco (cols 24-29)
                '1234567890',                 # Numero CCSS
                'Si',                         # Asegurado CCSS
                'CR65015200000000000000',      # Cuenta Bancaria / IBAN
                '88887777',                   # SINPE Movil
                'BNCR',                       # Banco
                'Cuenta Corriente',           # Tipo de Cuenta Banco
                # -- Configuracion nomina (cols 30-34)
                '750000',                     # Salario Base (CRC)
                '15/03/2021',                 # Fecha Vigencia Salarial
                'Salario Fijo',               # Metodo de Calculo
                'Costarricense',              # Nacionalidad INS
                'No',                         # Salario Variable
                # -- Datos personales y medicos (cols 35-43)
                'Soltero/a',                  # Estado Civil INS
                'Masculino',                  # Genero
                'Costa Rica',                 # Pais
                '1',                          # Numero de Dependientes
                'San Jose, Escazu, Res. Los Laureles',  # Direccion
                '88990011',                   # Telefono Personal
                '',                           # Correo Personal
                '',                           # Diagnostico / Notas Medicas
                'WARN FILA DE PRUEBA -- eliminar antes de importar',  # Observaciones
            ]

        self._build_rows(ws, cols, data_rows=100, header_row=3, example_row=4,
                         sample_values=sv)

        # -- Dropdowns estaticos -------------------------------------------
        self._dv(ws,  3, 'id_type',        5, title='Tipo de Identificacion')
        # col 13: Tipo de Horario -- DINAMICO (lee planilla.schedule.type de BD)
        # (se aplica abajo junto con los demas dinamicos)
        self._dv(ws, 15, 'si_no',          5, title='Incluir en INS (si/no)')
        self._dv(ws, 20, 'ins_risk',       5, title='Clase de Riesgo INS')
        self._dv(ws, 21, 'ins_workday',    5, title='Tipo de Jornada INS')
        self._dv(ws, 22, 'ins_occupation', 5, title='Ocupacion INS (COCR-2023)')
        self._dv(ws, 23, 'blood_type',     5, title='Tipo de Sangre')
        self._dv(ws, 25, 'si_no',          5, title='Asegurado CCSS (si/no)')
        self._dv(ws, 28, 'banco',          5, title='Banco')
        self._dv(ws, 29, 'account_type',   5, title='Tipo de Cuenta Banco')
        # col 30: Salario Base -- sin dropdown
        # col 31: Fecha Vigencia -- sin dropdown
        self._dv(ws, 32, 'calc_method',    5, title='Metodo de Calculo')
        self._dv(ws, 33, 'ins_nat',        5, title='Nacionalidad INS')
        self._dv(ws, 34, 'si_no',          5, title='Salario Variable (si/no)')
        self._dv(ws, 35, 'ins_civil',      5, title='Estado Civil INS')
        self._dv(ws, 36, 'gender',         5, title='Genero')

        # -- Dropdowns dinamicos (desde BD) --------------------------------
        dl = dyn_lists or {}
        self._dv_dynamic(ws,  7, 'department',     5, dl, title='Departamento')
        self._dv_dynamic(ws,  8, 'subdepartment',  5, dl, title='Sub Departamento')
        self._dv_dynamic(ws,  9, 'branch',         5, dl, title='Sucursal')
        self._dv_dynamic(ws, 10, 'job',            5, dl, title='Puesto / Cargo')
        self._dv_dynamic(ws, 11, 'employee_type',  5, dl, title='Tipo de Empleado')
        self._dv_dynamic(ws, 12, 'employee_status',5, dl, title='Estado del Empleado')
        self._dv_dynamic(ws, 13, 'schedule',       5, dl, title='Tipo de Horario')
        self._dv_dynamic(ws, 14, 'calendar',       5, dl, title='Calendarizacion de Planilla')
        self._dv_dynamic(ws, 37, 'country',        5, dl, title='Pais')
    # HOJA PRESTAMOS
    def _build_loans(self, wb, sample=False):
        cols = [
            ('Cedula Empleado',          True,  18, '1-2345-6789',     'Debe coincidir con hoja EMPLEADOS'),
            ('Tipo de Prestamo',         True,  16, 'Prestamo de Empresa', 'Prestamo de Empresa / Adelanto de Salario'),
            ('Descripcion / Motivo',     False, 30, 'Prestamo personal',''),
            ('Monto Total (CRC)',          True,  16, '500000',          'Total del prestamo, sin comas'),
            ('Numero de Cuotas',         True,  14, '10',              'Cantidad de cuotas a descontar'),
            ('Fecha de Otorgamiento',    True,  18, '15/01/2026',      'DD/MM/AAAA'),
            ('Fecha Primera Deduccion',  True,  18, '01/02/2026',      'DD/MM/AAAA -- primer boleta que descuenta'),
            ('Estado',                   True,  14, 'Aprobado',            'Ver CATALOGOS -> loan_state'),
            ('Monto ya Pagado (CRC)',      False, 16, '100000',          'Si ya se ha descontado algo'),
            ('Observaciones',            False, 28, '',                ''),
        ]
        sv = [self._SAMPLE_CEDULA, 'Prestamo de Empresa', 'Prestamo de prueba', '100000', '5',
              '01/01/2024', '01/02/2024', 'Aprobado', '0', 'WARN PRUEBA'] if sample else None
        ws = wb.create_sheet(' PRESTAMOS')
        self._sheet_title(ws, 'PRESTAMOS Y ADELANTOS -- Un prestamo por fila (puede haber varios por empleado)', len(cols))
        self._build_rows(ws, cols, data_rows=100, sample_values=sv)
        # col 2: Tipo de Prestamo, col 8: Estado
        self._dv(ws, 2, 'loan_type',   4, title='Tipo de Prestamo')
        self._dv(ws, 8, 'loan_state',  4, title='Estado del Prestamo')

    # ==========================================================================
    # HOJA PENSION ALIMENTARIA
    # ==========================================================================
    def _build_pension(self, wb, sample=False):
        cols = [
            ('Cedula Empleado',        True,  18, '1-2345-6789',          'Cedula del empleado afectado'),
            ('Numero de Expediente',   True,  22, '15-000123-0638-FA',    'Numero del expediente judicial'),
            ('Juzgado',                True,  30, 'Juzgado de Familia SJ', ''),
            ('Fecha de Resolucion',    True,  18, '10/06/2023',            'DD/MM/AAAA'),
            ('Nombre Beneficiario',    True,  26, 'Maria Rodriguez Solano','Nombre completo'),
            ('Relacion Beneficiario',  True,  16, 'Hijo/a',                  'Ver CATALOGOS -> pension_relacion'),
            ('Cuenta Beneficiario',    False, 28, 'CR21015108010018023571','IBAN del beneficiario (opcional)'),
            ('Tipo de Calculo',        True,  16, 'Porcentaje del Salario', 'Porcentaje del Salario / Monto Fijo'),
            ('Porcentaje (%)',          False, 12, '25',                   'Si tipo=porcentaje, solo el numero (ej: 25)'),
            ('Monto Fijo (CRC)',         False, 14, '',                     'Si tipo=monto_fijo, monto en colones'),
            ('Fecha de Inicio',        True,  14, '01/07/2023',            'DD/MM/AAAA'),
            ('Fecha de Fin',           False, 14, '',                     'Dejar vacio si no tiene vencimiento'),
        ]
        sv = [self._SAMPLE_CEDULA, 'TEST-0000-PRUEBA', 'Juzgado Prueba', '01/01/2024',
              'Beneficiario Prueba', 'Hijo/a', '', 'Porcentaje del Salario', '10', '',
              '01/01/2024', ''] if sample else None
        ws = wb.create_sheet(' PENSION_ALIMENTARIA')
        self._sheet_title(ws, 'PENSIONES ALIMENTARIAS -- Una resolucion por fila', len(cols))
        self._build_rows(ws, cols, sample_values=sv)
        # col 6: Relacion Beneficiario, col 8: Tipo de Calculo
        self._dv(ws, 6, 'pension_rel',  4, title='Relacion Beneficiario')
        self._dv(ws, 8, 'pension_calc', 4, title='Tipo de Calculo')

    # ==========================================================================
    # HOJA BENEFICIOS RECURRENTES
    # ==========================================================================
    def _build_benefits(self, wb, sample=False):
        cols = [
            ('Cedula Empleado',   True,  18, '1-2345-6789',        'Cedula del empleado'),
            ('Concepto',          True,  28, 'Cuota Sindical',      'Nombre descriptivo del descuento o deduccion'),
            ('Tipo',              True,  14, 'Deduccion / Descuento','Deduccion / Descuento   o   Beneficio / Ingreso'),
            ('Tipo de Monto',     True,  16, 'Monto Fijo',           'Monto Fijo / Porcentaje'),
            ('Monto (CRC)',         False, 14, '15000',               'Si tipo_monto=fijo'),
            ('Porcentaje (%)',    False, 12, '',                    'Si tipo_monto=porcentaje, solo el numero'),
            ('Codigo Deduccion',  False, 16, '',                    'Codigo del concepto si el modulo lo requiere'),
            ('Vigente Desde',     True,  14, '01/01/2026',          'DD/MM/AAAA'),
            ('Vigente Hasta',     False, 14, '',                    'Dejar vacio si es indefinido'),
            ('Nota',              False, 28, 'Acuerdo colectivo 2026','Descripcion o referencia'),
        ]
        sv = [self._SAMPLE_CEDULA, 'Cuota Sindical Prueba', 'Deduccion / Descuento', 'Monto Fijo',
              '2000', '', '', '01/01/2024', '', 'WARN PRUEBA'] if sample else None
        ws = wb.create_sheet(' OTROS DESCUENTOS')
        self._sheet_title(ws, 'OTROS DESCUENTOS / DEDUCCIONES RECURRENTES -- Cuota sindical, cooperativa, ahorro voluntario, seguro medico, etc.', len(cols))
        self._build_rows(ws, cols, data_rows=100, sample_values=sv)
        # col 3: Tipo, col 4: Tipo de Monto
        self._dv(ws, 3, 'benefit_type', 4, title='Tipo (beneficio/deduccion)')
        self._dv(ws, 4, 'amount_type',  4, title='Tipo de Monto')

    def _build_disabilities(self, wb, sample=False):
        cols = [
            ('Cedula Empleado',           True,  18, '1-2345-6789',         ''),
            ('Tipo de Incapacidad',       True,  22, 'Enfermedad Comun (CCSS)', 'Ver CATALOGOS -> disability_type'),
            ('Fecha Inicio',              True,  14, '01/02/2026',           'DD/MM/AAAA'),
            ('Fecha Fin',                 True,  14, '10/02/2026',           'DD/MM/AAAA'),
            ('% Subsidiado CCSS',         False, 14, '60',                   'Porcentaje que paga la CCSS (60%% en enfermedad, 100%% en maternidad)'),
            ('% a Cargo Patrono',         False, 14, '0',                    'Porcentaje que asume el patrono (0 por defecto salvo convenio)'),
            ('Numero Certificado',        False, 20, 'CCSS-2026-123',        'Numero del certificado CCSS o INS'),
            ('Diagnostico',               False, 28, 'Gripa severa',         'Descripcion del diagnostico'),
            ('Salario Diario (CRC)',       False, 16, '25000',                'Salario mensual / 30'),
            # -- Campos especiales Maternidad ----------------------------
            ('Fecha de Parto',            False, 14, '',                     'Solo maternidad: DD/MM/AAAA'),
            ('Maternidad 50/50 (Si/No)',  False, 14, 'Si',                   'Maternidad: Si=Patrono 50%% + CCSS 50%%  |  No=CCSS 100%% (Art. 94 CT)'),
            ('Cobrar CCSS obrera 10.83%% s/patronal (Si/No)', False, 14, 'No', 'Maternidad 50/50: Si=aplica 10.83%% CCSS obrera sobre el 50%% que paga el patrono'),
            ('Observaciones',             False, 28, '',                     ''),
        ]
        sv = [self._SAMPLE_CEDULA, 'Maternidad / Paternidad', '16/03/2026', '13/07/2026',
              '100', '0', 'CCSS-MAT-001', 'Licencia maternidad', '18333',
              '15/04/2026', 'Si', 'No', 'WARN PRUEBA'] if sample else None
        # Nota: date_start = parto(15/04) - 30 dias = 16/03 (prenatal valido)
        #       date_end  = parto(15/04) + 90 dias = 13/07 (postnatal valido)
        #       Total: 120 dias exactos (maximo legal Art. 94 CT)
        ws = wb.create_sheet(' INCAPACIDADES')
        self._sheet_title(ws, 'INCAPACIDADES -- Solo las activas o dentro del periodo de carga', len(cols))
        self._build_rows(ws, cols, sample_values=sv)
        # col 2: Tipo de Incapacidad
        self._dv(ws, 2, 'disability', 4, title='Tipo de Incapacidad')
        # col 11: Maternidad 50/50, col 12: Cobrar CCSS
        self._dv(ws, 11, 'si_no', 4, title='Maternidad 50/50')
        self._dv(ws, 12, 'si_no', 4, title='Cobrar CCSS sobre patronal')

    def _build_vacations(self, wb, sample=False):
        cols = [
            # -- Identificacion ------------------------------------------------
            ('Cedula Empleado',               True,  18, '1-2345-6789',   'Cedula del empleado tal como esta en el sistema'),
            # -- Saldo inicial pre-implementacion -----------------------------
            ('Saldo Inicial (dias)',           True,  18, '8.50',
             'OBLIGATORIO: dias de vacaciones disponibles a la Fecha de Corte.\n'
             'Es el saldo REAL que tiene el empleado hoy (lo que le falta por disfrutar).\n'
             'Ejemplo: si tiene 8.5 dias pendientes, escriba 8.5'),
            ('Fecha de Corte del Saldo',      True,  20, '31/12/2025',
             'OBLIGATORIO: fecha exacta hasta la cual se calculo el saldo inicial.\n'
             'El sistema acumulara dias solo a partir de esta fecha.\n'
             'Use el ultimo dia antes de que empiece a usar el sistema.\n'
             'Formato: DD/MM/AAAA'),
            # -- Informacion de referencia (solo para documentacion) -----------
            ('Dias Acumulados Totales (ref)', False, 20, '24.00',
             'Referencia: total de dias que le correspondian desde su ingreso.\n'
             'No se importa, solo para documentar el calculo.'),
            ('Dias Tomados Historial (ref)',  False, 20, '15.50',
             'Referencia: dias que ya disfruto antes de la implementacion.\n'
             'No se importa, solo para documentar el calculo.\n'
             'Verificacion: Acumulados  Tomados = Saldo Inicial'),
            ('Salario Diario Ref. (CRC)',       False, 20, '25000',
             'Referencia: salario diario del empleado a la fecha de corte.\n'
             'No se importa, solo para documentar cuanto valdria cada dia.'),
            ('Observaciones',                 False, 30, 'Saldo calculado por RRHH a dic-2025',
             'Notas internas, quien calculo el saldo, fuente del dato, etc.'),
        ]
        sv = [self._SAMPLE_CEDULA, '5.0', '31/12/2025', '17.0', '12.0',
              '16667', 'WARN PRUEBA'] if sample else None
        ws = wb.create_sheet(' VACACIONES')
        self._sheet_title(
            ws,
            'SALDO INICIAL DE VACACIONES -- Para empresas con empleados pre-existentes',
            len(cols)
        )
        self._build_rows(ws, cols, sample_values=sv)

        # Instrucciones adicionales al final de la hoja
        last_row = ws.max_row + 2
        inst_fill = PatternFill('solid', fgColor='EBF5FB')
        inst_font = Font(name='Calibri', size=10, italic=True, color='1A5276')

        instructions = [
            '  INSTRUCCIONES DE USO:',
            '',
            '  1. Complete CEDULA + SALDO INICIAL + FECHA DE CORTE para cada empleado.',
            '  2. El "Saldo Inicial" es la cantidad de dias disponibles en esa fecha exacta.',
            '     Ejemplo: Juan tiene derecho a 24 dias y ha tomado 15.5 -> Saldo = 8.5 dias.',
            '  3. La "Fecha de Corte" debe ser el dia anterior a que el sistema empiece a controlar.',
            '     Ejemplo: si arranca en Enero 2026 -> use 31/12/2025.',
            '  4. A partir de esa fecha el sistema acumulara dias nuevos automaticamente.',
            '  5. Las columnas "Dias Acumulados" y "Dias Tomados" son solo para documentar --',
            '     NO afectan la importacion. Use la columna de verificacion: Acumulados  Tomados = Saldo.',
            '  6. Si el empleado NO tiene saldo inicial (ingreso despues del sistema), deje en 0.',
        ]
        for i, line in enumerate(instructions):
            cell = ws.cell(row=last_row + i, column=1, value=line)
            cell.fill = inst_fill
            cell.font = inst_font
            ws.merge_cells(
                start_row=last_row + i, start_column=1,
                end_row=last_row + i,   end_column=len(cols)
            )


    def _build_overtime(self, wb, sample=False):
        cols = [
            ('Cedula Empleado',      True,  18, '1-2345-6789',    ''),
            ('Fecha',                True,  14, '01/02/2026',      'DD/MM/AAAA'),
            ('Tipo de Hora Extra',   True,  20, 'Simple (1.5x)',        'Ver CATALOGOS -> overtime_type'),
            ('Cantidad de Horas',    True,  16, '2.5',             'Numero de horas extras trabajadas'),
            ('Salario por Hora (CRC)', False, 18, '3500',            'Salario mensual / 240 (o segun contrato)'),
            ('Monto Total (CRC)',      False, 16, '8750',            'Horas x Salario x Factor (1.5 / 2.0)'),
            ('Periodo de Planilla',  False, 22, 'Febrero 2026',    'Periodo al que se carga esta hora extra'),
            ('Observaciones',        False, 28, '',                 ''),
        ]
        sv = [self._SAMPLE_CEDULA, '15/01/2024', 'Simple (1.5x)', '2',
              '2083', '6250', 'Enero 2024', 'WARN PRUEBA'] if sample else None
        ws = wb.create_sheet(' HORAS EXTRAS')
        self._sheet_title(ws, 'HORAS EXTRAS -- Registros historicos a importar', len(cols))
        self._build_rows(ws, cols, sample_values=sv)
        # col 3: Tipo de Hora Extra
        self._dv(ws, 3, 'overtime_type', 4, title='Tipo de Hora Extra')

    # ==========================================================================
    # HOJA OCULTA DE LISTAS (fuente de los dropdowns)
    # ==========================================================================
    # ==========================================================================
    # HOJA EMBARGOS JUDICIALES
    # ==========================================================================
    def _build_embargos(self, wb, sample=False):
        cols = [
            ('Cedula Empleado',        True,  18, '1-2345-6789',          'Cedula del empleado afectado'),
            ('Ndeg Expediente Judicial', True,  24, '15-000456-0638-CI',    'Numero del expediente del juzgado'),
            ('Juzgado / Tribunal',     True,  30, 'Juzgado Civil SJ',     'Nombre completo del juzgado'),
            ('Fecha de Resolucion',    False, 16, '15/01/2024',           'DD/MM/AAAA'),
            ('Nombre del Acreedor',    True,  28, 'Empresa XYZ S.A.',     'Nombre del beneficiario del embargo'),
            ('IBAN del Acreedor',      False, 30, 'CR21015108010018023571','IBAN para girar el embargo (opcional)'),
            ('Tipo de Calculo',        True,  22, 'Monto Fijo',           'Ver CATALOGOS -> embargo_calc'),
            ('Monto Fijo (CRC)',         False, 16, '50000',                'Si tipo = Monto Fijo'),
            ('Porcentaje (%)',          False, 12, '',                     'Si tipo = Porcentaje. Max 25% (Art. 172 CT)'),
            ('Vigente Desde',          True,  14, '01/02/2024',           'DD/MM/AAAA'),
            ('Vigente Hasta',          False, 14, '',                     'Dejar vacio si no tiene vencimiento'),
            ('Observaciones',          False, 28, '',                     ''),
        ]
        sv = [self._SAMPLE_CEDULA, 'TEST-EMB-0000', 'Juzgado Prueba', '01/01/2024',
              'Acreedor Prueba', '', 'Monto Fijo', '10000', '',
              '01/01/2024', '', 'WARN PRUEBA'] if sample else None
        ws = wb.create_sheet(' EMBARGOS')
        self._sheet_title(ws, 'EMBARGOS JUDICIALES -- Art. 172 CT (max. 25% del neto disponible)', len(cols))
        self._build_rows(ws, cols, data_rows=80, sample_values=sv)
        self._dv(ws, 7, 'embargo_calc', 4, title='Tipo de Calculo')

    # ==========================================================================
    # HOJA BONOS E INCENTIVOS
    # ==========================================================================
    def _build_bonos(self, wb, sample=False):
        cols = [
            ('Cedula Empleado',        True,  18, '1-2345-6789',                    'Cedula del empleado'),
            ('Concepto / Nombre',      True,  28, 'Bono de Productividad Q1 2024',  'Nombre descriptivo del bono'),
            ('Tipo de Bono',           True,  28, 'Productividad / Rendimiento',    'Ver CATALOGOS -> bono_type'),
            ('Tipo de Calculo',        True,  18, 'Monto Fijo',                     'Monto Fijo / Porcentaje del Salario Base'),
            ('Monto Fijo (CRC)',         False, 16, '25000',                          'Si tipo calculo = Monto Fijo'),
            ('Porcentaje (%)',          False, 12, '',                               'Si tipo calculo = Porcentaje'),
            ('Es Recurrente',          True,  14, 'Si',                             'Si = se aplica cada boleta / No = solo una vez'),
            ('Afecto CCSS',            True,  12, 'Si',                             'Si = suma a base CCSS (bonos salariales)'),
            ('Afecto Renta',           True,  12, 'Si',                             'Si = suma a base de renta'),
            ('Tope Exento (CRC/mes)',    False, 16, '',                               'Solo para transporte (CRC74 000/mes) o similar'),
            ('Vigente Desde',          True,  14, '01/01/2024',                     'DD/MM/AAAA'),
            ('Vigente Hasta',          False, 14, '',                               'Dejar vacio para aplicar indefinidamente'),
            ('Observaciones',          False, 30, 'Acuerdo de junta 2024',          ''),
        ]
        sv = [self._SAMPLE_CEDULA, 'Bono Prueba', 'Productividad / Rendimiento',
              'Monto Fijo', '5000', '', 'Si', 'Si', 'Si', '',
              '01/01/2024', '', 'WARN PRUEBA'] if sample else None
        ws = wb.create_sheet(' BONOS')
        self._sheet_title(ws, 'BONOS E INCENTIVOS -- Aplican automaticamente en cada boleta', len(cols))
        self._build_rows(ws, cols, data_rows=100, sample_values=sv)
        self._dv(ws, 3, 'bono_type',        4, title='Tipo de Bono')
        self._dv(ws, 4, 'bono_calc',        4, title='Tipo de Calculo')
        self._dv(ws, 7, 'si_no_recurrente', 4, title='Es Recurrente')
        self._dv(ws, 8, 'si_no',            4, title='Afecto CCSS')
        self._dv(ws, 9, 'si_no',            4, title='Afecto Renta')

    # ==========================================================================
    # HOJA CATALOGOS
    # ==========================================================================
    # ==========================================================================
    # HOJA ACUMULADOS DE PROVISIONES (aguinaldo, cesantia, vacaciones)
    # ==========================================================================
    def _build_acumulados(self, wb, sample=False):
        """
        Hoja para cargar los acumulados de provisiones de empleados pre-existentes.
        Necesario para instalaciones nuevas donde la empresa ya tiene empleados
        con tiempo laborado y provisiones acumuladas historicas.
        """
        cols = [
            # -- Identificacion -----------------------------------------------
            ('Cedula Empleado',              True,  18, '1-2345-6789',
             'Cedula del empleado (llave con hoja EMPLEADOS)'),
            ('Nombre Empleado (ref)',         False, 28, 'Juan Perez Rodriguez',
             'Referencia -- no se importa'),
            ('Fecha de Corte',               True,  16, '31/12/2025',
             'Fecha hasta la cual estan calculados los acumulados.\n'
             'Debe ser el dia anterior al inicio del sistema.\n'
             'Formato: DD/MM/AAAA'),
            # -- Aguinaldo ----------------------------------------------------
            ('Aguinaldo Acumulado (CRC)',     True,  20, '125000',
             'Monto en colones acumulado de aguinaldo hasta la fecha de corte.\n'
             'Calculo manual: (Salario_Mensual / 12) * meses_trabajados_en_el_anho.\n'
             'Ejemplo: empleado con 8 meses laborados en 2025 y salario 600,000:\n'
             '  600,000 / 12 * 8 = CRC 400,000'),
            ('Meses Laborados 2025 (ref)',    False, 18, '8',
             'Referencia: meses trabajados en el ano actual para calcular aguinaldo.\n'
             'No se importa.'),
            # -- Cesantia -----------------------------------------------------
            ('Cesantia Acumulada (CRC)',      True,  20, '98500',
             'Provision de cesantia acumulada hasta la fecha de corte.\n'
             'Calculo segun tabla Art. 29 CT por anos de servicio:\n'
             '  < 1 ano: 5.4167%% | 1 ano: 5.5556%% | ... | 7+ anos: 6.3889%%\n'
             'Aplica sobre cada salario pagado desde el ingreso.\n'
             'Use: Salario_Mensual * tasa_cesantia * meses_totales'),
            ('Anos de Servicio (ref)',        False, 14, '2.5',
             'Referencia: anos laborados a la fecha de corte. No se importa.\n'
             'Sirve para calcular la tasa de cesantia correcta.'),
            # -- Vacaciones ---------------------------------------------------
            ('Vacaciones Acumuladas (dias)',  True,  22, '8.5',
             'Dias de vacaciones disponibles a la fecha de corte.\n'
             'Ver tambien la hoja VACACIONES para mas detalle.\n'
             'Ejemplo: 8.5 dias pendientes de disfrutar.'),
            ('Valor Dia Vacacion (CRC)',      False, 18, '20000',
             'Referencia: valor de cada dia de vacaciones (Salario_Mensual / 30).\n'
             'No se importa.'),
            ('Vacaciones en Colones (ref)',   False, 18, '170000',
             'Referencia: dias * valor_dia. No se importa.'),
            # -- Provision total ----------------------------------------------
            ('Total Provisiones (CRC)',       False, 20, '393500',
             'Suma: Aguinaldo + Cesantia + (Vacaciones en colones).\n'
             'Solo referencia contable -- no se importa.'),
            ('Observaciones',                False, 36, 'Calculado por RRHH a dic-2025',
             'Quien calculo los acumulados, fuente del dato, etc.'),
        ]
        sv = [
            self._SAMPLE_CEDULA, 'Empleado Prueba', '31/12/2025',
            '400000', '8',
            '98500', '2.5',
            '8.5', '20000', '170000',
            '668500', 'WARN PRUEBA'
        ] if sample else None

        ws = wb.create_sheet(' ACUMULADOS')
        self._sheet_title(
            ws,
            'ACUMULADOS DE PROVISIONES -- Para instalaciones nuevas con empleados pre-existentes',
            len(cols)
        )
        self._build_rows(ws, cols, sample_values=sv)

        # Instrucciones al pie
        last_row = ws.max_row + 2
        inst_fill = PatternFill('solid', fgColor='EBF5FB')
        inst_font = Font(name='Calibri', size=10, italic=True, color='1A5276')
        instructions = [
            '  INSTRUCCIONES -- HOJA ACUMULADOS DE PROVISIONES',
            '',
            '  Esta hoja es para instalaciones NUEVAS donde la empresa ya tiene empleados con tiempo laborado.',
            '  Permite cargar los acumulados historicos de aguinaldo, cesantia y vacaciones',
            '  para que el sistema refleje el pasivo real desde el dia 1.',
            '',
            '  AGUINALDO (Art. 228 CT):',
            '    Calculo: Salario_Mensual / 12 * meses_laborados_en_el_anho_actual',
            '    Ejemplo: salario CRC 600,000, 8 meses en 2025 -> 600,000/12*8 = CRC 400,000',
            '',
            '  CESANTIA (Art. 29 CT -- tabla por anos de servicio):',
            '    < 1 ano:  5.4167%%  | 1 ano: 5.5556%%  | 2 anos: 5.6944%%',
            '    3 anos:  5.8333%%  | 4 anos: 5.9722%%  | 5 anos: 6.1111%%',
            '    6 anos:  6.2500%%  | 7+ anos: 6.3889%% (maximo legal)',
            '    Calculo: Salario_Mensual * tasa * meses_totales',
            '',
            '  VACACIONES (Art. 153 CT):',
            '    4.1667%% del salario mensual por cada mes trabajado = 0.5 dias por mes laborado',
            '    Saldo = dias_acumulados - dias_disfrutados',
            '',
            '  NOTA: Esta hoja es de referencia -- el sistema usa las hojas VACACIONES para',
            '  el saldo de dias. Los acumulados en colones son para informacion contable.',
        ]
        for i, line in enumerate(instructions):
            cell = ws.cell(row=last_row + i, column=1, value=line)
            cell.fill = inst_fill
            cell.font = inst_font
            ws.merge_cells(
                start_row=last_row + i, start_column=1,
                end_row=last_row + i,   end_column=len(cols)
            )

    # ==========================================================================
    # HOJA CONFIGURACION INICIAL
    # ==========================================================================
    def _build_config_inicial(self, wb):
        """
        Hoja de configuracion que recuerda al administrador los pasos clave
        para dejar el sistema listo: tramos de renta, calendarizacion por defecto,
        configuracion de CCSS, etc.
        """
        ws = wb.create_sheet(' CONFIGURACION')
        ws.title = ' CONFIGURACION'

        # Titulo principal
        titulo_fill = PatternFill('solid', fgColor='1B2631')
        titulo_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        ws.merge_cells('A1:F1')
        c = ws['A1']
        c.value = 'LISTA DE VERIFICACION -- Configuracion Inicial del Sistema Planilla CR'
        c.fill = titulo_fill; c.font = titulo_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        rows = [
            # (seccion, item, descripcion, ruta en odoo, estado)
            ('PASO 1 -- TRAMOS DE RENTA', '', '', '', ''),
            ('', 'Tramos DGT-R-016-2026 activos',
             'Verificar que haya 5 tramos activos para 2026',
             'Planilla -> Configuracion -> Tramos de Renta',
             'PENDIENTE'),
            ('', 'Tramo exento correcto',
             'El tramo exento debe ser CRC 918,000 (no 941,000)',
             'Editar el tramo exento y verificar el limite',
             'PENDIENTE'),
            ('', 'Tramos sin empresa asignada (globales)',
             'Para multi-empresa: dejar company_id en blanco para compartir tramos',
             'Editar cada tramo -> campo Empresa -> borrar',
             'PENDIENTE'),
            ('', '', '', '', ''),
            ('PASO 2 -- CALENDARIZACION POR DEFECTO', '', '', '', ''),
            ('', 'Configurar calendarizacion por defecto',
             'Seleccionar la calendarizacion que mas se usa (ej: Quincenal)',
             'Planilla -> Configuracion -> Config. Contable -> Calendarizacion por Defecto',
             'PENDIENTE'),
            ('', 'Crear calendarizaciones necesarias',
             'Crear las calendarizaciones de pago de su empresa',
             'Planilla -> Configuracion -> Calendarizaciones',
             'PENDIENTE'),
            ('', '', '', '', ''),
            ('PASO 3 -- CONFIGURACION CCSS', '', '', '', ''),
            ('', 'Tasas CCSS 2026',
             'CCSS obrero: 10.83%% | CCSS patronal: 26.83%%',
             'Planilla -> Configuracion -> Codigos de Deduccion -> CCSS_OBR',
             'PENDIENTE'),
            ('', 'INS clase de riesgo por defecto',
             'Verificar que los empleados tienen asignada la clase de riesgo correcta',
             'RR.HH -> Empleados -> tab Planilla CR -> Clase de Riesgo INS',
             'PENDIENTE'),
            ('', '', '', '', ''),
            ('PASO 4 -- IMPORTAR EMPLEADOS', '', '', '', ''),
            ('', 'Importar hoja EMPLEADOS',
             'Cargar empleados con cedula, salario, fecha ingreso, calendarizacion',
             'Planilla -> Importar Datos -> Empleados',
             'PENDIENTE'),
            ('', 'Importar hoja VACACIONES',
             'Cargar saldos de vacaciones historicos (fecha de corte)',
             'Planilla -> Importar Datos -> Saldos Vacaciones',
             'PENDIENTE'),
            ('', 'Importar hoja ACUMULADOS',
             'Registrar acumulados de aguinaldo, cesantia y vacaciones en colones',
             'Referencia contable -- cargar manualmente si aplica',
             'PENDIENTE'),
            ('', '', '', '', ''),
            ('PASO 5 -- PRIMERA PLANILLA', '', '', '', ''),
            ('', 'Crear planilla prueba',
             'Crear planilla en borrador para verificar calculos antes de confirmar',
             'Planilla -> Planillas -> Nuevo',
             'PENDIENTE'),
            ('', 'Verificar tramos de renta',
             'Confirmar que la renta se calcula con los tramos 2026 (exento 918k)',
             'Abrir una boleta -> verificar Resumen Completo -> Impuesto de Renta',
             'PENDIENTE'),
            ('', 'Verificar cesantia por anos de servicio',
             'Empleado con 5 anos debe usar tasa 5.9722%% (no 5.33%% fijo)',
             'Abrir una boleta -> Cargas Patronales -> Provision Cesantia',
             'PENDIENTE'),
        ]

        # Headers de columnas
        headers = ['Seccion / Paso', 'Item de Verificacion', 'Descripcion',
                   'Ruta en Odoo', 'Estado']
        widths   = [32, 42, 55, 55, 14]
        hdr_fill = PatternFill('solid', fgColor='1A5276')
        hdr_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=2, column=col, value=h)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[2].height = 18

        # Colores por tipo de fila
        seccion_fill = PatternFill('solid', fgColor='2E4053')
        seccion_font = Font(name='Calibri', size=10, bold=True, color='F9E79F')
        item_font    = Font(name='Calibri', size=10)
        alt_fill     = PatternFill('solid', fgColor='EBF5FB')
        status_fill  = {
            'PENDIENTE': PatternFill('solid', fgColor='FDEBD0'),
            'LISTO':     PatternFill('solid', fgColor='D5F5E3'),
        }

        for i, row_data in enumerate(rows, 3):
            seccion, item, desc, ruta, estado = row_data
            is_seccion = bool(seccion and not item)
            for col, val in enumerate([seccion or item, item if seccion else '', desc, ruta, estado], 1):
                c = ws.cell(row=i, column=col, value=val)
                if is_seccion:
                    c.fill = seccion_fill; c.font = seccion_font
                else:
                    c.font = item_font
                    if i % 2 == 0: c.fill = alt_fill
                c.alignment = Alignment(wrap_text=True, vertical='center')
            # Status cell dropdown
            if estado in ('PENDIENTE', 'LISTO'):
                dv = DataValidation(type='list', formula1='"PENDIENTE,LISTO,NO APLICA"', showDropDown=False)
                ws.add_data_validation(dv)
                dv.add(ws.cell(row=i, column=5))
                st_cell = ws.cell(row=i, column=5)
                st_cell.fill = status_fill.get(estado, PatternFill())
            ws.row_dimensions[i].height = 30


    def _build_catalogs(self, wb):
        C = self._C
        ws = wb.create_sheet(' CATALOGOS')
        ws.sheet_view.showGridLines = False

        ws.merge_cells('A1:C1')
        c = ws['A1']
        c.value     = 'CATALOGOS DE VALORES VALIDOS -- WARN No editar esta hoja'
        c.font      = self._font(bold=True, color=C['white'], size=12)
        c.fill      = self._fill(C['red_hdr'])
        c.alignment = self._center()
        ws.row_dimensions[1].height = 28

        CATALOGS = [
            ('id_type -- Tipo de Identificacion', [
                ('01',  'Cedula Nacional'),
                ('02',  'Residencia / DIMEX'),
                ('03',  'Permiso de Trabajo'),
                ('04',  'Pasaporte'),
                ('05',  'Indocumentado'),
            ]),
            ('employee_type -- Tipo de Empleado (buscar por nombre exacto en Odoo)', [
                ('planilla',    'Ejemplo: nombre del tipo tal como aparece en Configuracion -> Tipos de Empleado'),
                ('contratado',  'Use el nombre exacto del registro en Odoo'),
            ]),
            ('employee_status -- Estado del Empleado (buscar por nombre en Odoo)', [
                ('activo',      'Use el nombre exacto del estado en Configuracion -> Estados'),
            ]),
            ('schedule_type -- Tipo de Horario (nombre exacto en Odoo)', [
                ('',  'Use el nombre del horario tal como aparece en Configuracion -> Tipos de Horario'),
            ]),
            ('frequency -- Frecuencia de Pago', [
                ('Mensual',    'Mensual -- 1 pago al mes'),
                ('Quincenal',  'Quincenal -- 2 pagos al mes'),
                ('Semanal',    'Semanal -- 4 pagos al mes'),
                ('Bimensual',  'Bimensual -- cada 2 meses'),
            ]),
            ('ins_risk_class -- Clase de Riesgo INS', [
                ('I',   'Clase I   -- Oficinas y administrativo (~0.87%)'),
                ('II',  'Clase II  -- Comercio (~1.49%)'),
                ('III', 'Clase III -- Industria liviana (~2.47%)'),
                ('IV',  'Clase IV  -- Construccion / riesgo alto (~4.13%)'),
                ('V',   'Clase V   -- Actividades de alto riesgo (~6.88%)'),
            ]),
            ('ins_workday_type -- Tipo de Jornada INS', [
                ('Ordinaria',      'Jornada diurna regular'),
                ('Extraordinaria', 'Horas extra autorizadas'),
                ('Mixta',          'Parte diurna y parte nocturna'),
                ('Tiempo Parcial', 'Menos de jornada completa'),
                ('Por Horas',      'Segun horas efectivamente trabajadas'),
                ('Ocasional',      'Trabajo esporadico o temporal'),
            ]),
            ('ins_id_type -- Tipo de ID INS', [
                ('01', 'Cedula de Costa Rica'),
                ('02', 'Residencia de Costa Rica / DIMEX'),
                ('03', 'Permiso de Trabajo'),
                ('04', 'Pasaporte'),
                ('05', 'Indocumentado'),
            ]),
            ('ins_civil_status -- Estado Civil INS', [
                ('Soltero/a',    ''),
                ('Casado/a',     ''),
                ('Divorciado/a', ''),
                ('Viudo/a',      ''),
                ('Union Libre',  ''),
                ('Separado/a',   ''),
            ]),
            ('ins_nationality -- Nacionalidad INS', [
                ('CR', 'Costarricense'),
                ('NI', 'Nicaraguense'),
                ('CO', 'Colombiana'),
                ('US', 'Estadounidense'),
                ('HN', 'Hondurena'),
                ('SV', 'Salvadorena'),
                ('GT', 'Guatemalteca'),
                ('PA', 'Panamena'),
                ('MX', 'Mexicana'),
                ('VE', 'Venezolana'),
                ('PE', 'Peruana'),
                ('EC', 'Ecuatoriana'),
                ('OT', 'Otra nacionalidad'),
            ]),
            ('account_type -- Tipo de Cuenta Banco', [
                ('Cuenta Corriente', 'Cuenta corriente o IBAN'),
                ('Cuenta de Ahorros','Cuenta de ahorros'),
                ('SINPE Movil',      'SINPE Movil'),
            ]),
            ('bank -- Banco', [
                ('BNCR',       'Banco Nacional de Costa Rica'),
                ('BCR',        'Banco de Costa Rica'),
                ('BP',         'Banco Popular y de Desarrollo Comunal'),
                ('BAC',        'BAC San Jose'),
                ('BCT',        'Banco BCT'),
                ('CATHAY',     'Banco Cathay de Costa Rica'),
                ('CMB',        'Banco CMB'),
                ('DAVIVIENDA', 'Banco Davivienda'),
                ('GENERAL',    'Banco General'),
                ('IMPROSA',    'Banco Improsa'),
                ('LAFISE',     'Banco La Fise'),
                ('PROMERICA',  'Banco Promerica'),
                ('PRIVAL',     'Prival Bank'),
                ('SCOTIA',     'Scotiabank'),
                ('COOCIQUE',   'Coocique R.L.'),
                ('COOPENAE',   'Coopenae R.L.'),
                ('MUTUAL_ALJ', 'Mutual Alajuela'),
                ('OTRO',       'Otro banco / cooperativa'),
            ]),
            ('loan_type -- Tipo de Prestamo', [
                ('Prestamo de Empresa', 'Prestamo otorgado por la empresa'),
                ('Adelanto de Salario', 'Adelanto sobre el salario del periodo'),
            ]),
            ('loan_state -- Estado del Prestamo', [
                ('Aprobado', 'Se activara en la proxima boleta'),
                ('En Curso', 'Descuento activo'),
                ('Borrador', 'Pendiente de aprobacion'),
                ('Pagado',   'Totalmente cancelado'),
                ('Anulado',  'Prestamo anulado'),
            ]),
            ('pension_relacion -- Relacion Beneficiario', [
                ('Hijo/a',   ''),
                ('Conyuge',  'Conyuge / Conviviente'),
                ('Padre',    ''),
                ('Madre',    ''),
                ('Otro',     ''),
            ]),
            ('pension_calc -- Tipo de Calculo Pension', [
                ('Porcentaje del Salario', 'Porcentaje del salario bruto'),
                ('Monto Fijo',             'Monto fijo mensual en colones'),
            ]),
            ('benefit_type -- Tipo de Descuento/Deduccion Recurrente', [
                ('Beneficio / Ingreso',    'Suma al salario bruto (ej: plus informal no cubierto por BONOS)'),
                ('Deduccion / Descuento',  'Resta del salario neto (ej: cuota sindical, cooperativa, ahorro)'),
            ]),
            ('amount_type -- Tipo de Monto', [
                ('Monto Fijo',  'Monto fijo en colones'),
                ('Porcentaje',  'Porcentaje del salario base'),
            ]),
            ('disability_type -- Tipo de Incapacidad', [
                ('Enfermedad Comun (CCSS)',     'Enfermedad o accidente no laboral'),
                ('Accidente de Trabajo (CCSS)', 'Accidente en el lugar de trabajo'),
                ('Riesgo Laboral (INS)',         'Cubierto por poliza INS'),
                ('Maternidad / Paternidad',      'Licencia pre/post natal'),
                ('Otro',                         'Otro tipo de incapacidad'),
            ]),
            ('overtime_type -- Tipo de Hora Extra', [
                ('Simple (1.5x)', 'Hora extra ordinaria -- factor 1.5'),
                ('Doble (2.0x)',  'Hora extra nocturna o dominical -- factor 2.0'),
                ('Dia Feriado',   'Trabajo en dia feriado nacional'),
            ]),
            ('embargo_calc -- Tipo de Calculo Embargo', [
                ('Monto Fijo',                  'Monto fijo en colones (CRC) cada periodo'),
                ('Porcentaje del Neto Disponible', 'Porcentaje del neto (bruto  CCSS  renta  pensiones). Max 25% Art. 172 CT'),
            ]),
            ('bono_type -- Tipo de Bono', [
                ('Productividad / Rendimiento',          'Afecto CCSS y Renta -- integra salario para aguinaldo/cesantia'),
                ('Asistencia Perfecta',                   'Afecto CCSS y Renta -- integra salario para aguinaldo/cesantia'),
                ('Antiguedad por Anos de Servicio',       'Afecto CCSS y Renta -- integra salario para aguinaldo/cesantia'),
                ('Subsidio de Transporte / Kilometraje',  'Exento CCSS/Renta hasta CRC74 000/mes (Reglamento 2023)'),
                ('Subsidio de Alimentacion (en dinero)',  'Afecto CCSS y Renta si se paga en dinero'),
                ('Subsidio Educativo',                    'Generalmente exento segun convenio colectivo'),
                ('Subsidio de Salud / Medico',            'Exento CCSS (Art. 5 Ley 7983) si es poliza medica'),
                ('Gastos de Representacion',              'Exento CCSS si estan debidamente documentados'),
                ('Comision por Ventas',                   'Afecto CCSS y Renta -- integra salario'),
                ('Incentivo / Premio Especial',           'Afecto CCSS y Renta'),
                ('Otro',                                  'Consulte con su contador el tratamiento fiscal'),
            ]),
            ('calc_method -- Metodo de Calculo de Planilla', [
                ('Salario Fijo',         'Sin consultar asistencias'),
                ('Por Horas Trabajadas', 'Segun modulo de asistencias'),
            ]),
        ]

        row = 3
        for titulo, valores in CATALOGS:
            # Encabezado de seccion
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row,   end_column=3)
            c = ws.cell(row, 1, value=titulo)
            c.font      = self._font(bold=True, color=C['white'], size=10)
            c.fill      = self._fill(C['med'])
            c.border    = self._border()
            c.alignment = self._left()
            for ci in (2, 3):
                ws.cell(row, ci).fill   = self._fill(C['med'])
                ws.cell(row, ci).border = self._border()
            ws.row_dimensions[row].height = 20
            row += 1

            # Sub-encabezado
            for ci, hdr in enumerate(['Valor a usar (exacto)', 'Descripcion'], 1):
                c = ws.cell(row, ci, value=hdr)
                c.font      = self._font(bold=True, color=C['dark'], size=9)
                c.fill      = self._fill(C['light'])
                c.border    = self._border()
                c.alignment = self._center()
            ws.row_dimensions[row].height = 16
            row += 1

            # Valores
            for val, desc in valores:
                cv = ws.cell(row, 1, value=val)
                cv.font      = self._font(bold=True, color='00008B', size=10)
                cv.fill      = self._fill(C['opt'])
                cv.border    = self._border()
                cv.alignment = self._left()

                cd = ws.cell(row, 2, value=desc)
                cd.font      = self._font(size=10)
                cd.fill      = self._fill(C['opt'])
                cd.border    = self._border()
                cd.alignment = self._left()
                ws.row_dimensions[row].height = 15
                row += 1

            row += 1  # separador

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 52
        ws.protection.sheet    = True
        ws.protection.password = 'planillacr2026'


    # ==========================================================================
    # HOJA COBROS A EMPLEADOS
    # ==========================================================================
    def _build_cobros_empleados(self, wb, sample=False):
        """
        Hoja para cargar cobros a empleados (almuerzos, productos, uniformes,
        multas, etc.) que se descuentan automaticamente en la boleta.
        """
        cols = [
            ('Cedula Empleado',         True,  18, '1-2345-6789',
             'Cedula del empleado (llave con hoja EMPLEADOS)'),
            ('Tipo de Cobro',           True,  28, 'Almuerzo',
             'Descripcion del tipo de cobro (ej: Almuerzo, Uniforme, Multa)'),
            ('Descripcion / Detalle',   False, 32, 'Almuerzo enero 2026',
             'Detalle adicional del cobro'),
            ('Precio Unitario (CRC)',    True,  18, '3500',
             'Monto por unidad o monto total si cantidad=1'),
            ('Cantidad',                True,  12, '1',
             'Numero de unidades. Para monto fijo usar 1'),
            ('Subsidio Patronal (%%)',   False, 18, '0',
             'Porcentaje que subsidia el patrono (0 si el empleado paga todo)'),
            ('Cobro Recurrente (Si/No)',True,  18, 'No',
             'Si=aplica cada periodo automaticamente | No=cobro de una sola vez'),
            ('Vigente Desde',           True,  14, '01/03/2026',
             'Fecha a partir de la cual aplica el cobro. Formato: DD/MM/AAAA'),
            ('Vigente Hasta',           False, 14, '',
             'Fecha de vencimiento. Dejar vacio si es indefinido o cobro unico.'),
            ('Afecta Base CCSS (Si/No)',False, 18, 'No',
             'Si=suma al salario cotizable CCSS | No=no afecta base (default)'),
            ('Estado',                  True,  14, 'Aprobado',
             'Borrador / Aprobado -- use Aprobado para que se aplique en la proxima boleta'),
            ('Observaciones',           False, 30, '',
             'Notas internas del cobro'),
        ]
        sv = [
            self._SAMPLE_CEDULA, 'Almuerzo', 'Almuerzo periodo prueba',
            '3500', '1', '0', 'No', '01/01/2026', '', 'No', 'Aprobado', 'WARN PRUEBA'
        ] if sample else None

        ws = wb.create_sheet(' COBROS_EMPLEADOS')
        self._sheet_title(
            ws,
            'COBROS A EMPLEADOS -- Almuerzos, uniformes, productos, multas, etc. (se descuentan en boleta)',
            len(cols)
        )
        self._build_rows(ws, cols, data_rows=100, sample_values=sv)
        self._dv(ws,  7, 'si_no', 4, title='Cobro Recurrente')
        self._dv(ws, 10, 'si_no', 4, title='Afecta Base CCSS')

    # ==========================================================================
    # ACCION PRINCIPAL
    # ==========================================================================
    def action_generate(self):
        self.ensure_one()

        wb = Workbook()
        s = self.include_sample_data

        # Hoja de listas -- debe crearse ANTES de las demas hojas
        _, static_cols = self._build_listas_sheet(wb)

        # Listas dinamicas desde la BD (catalogos que varian por empresa)
        dyn_lists = self._build_dynamic_lists(wb, self.company_id, static_cols)

        # Instrucciones siempre presentes
        self._build_instructions(wb)

        if self.include_employees:
            self._build_employees(wb, sample=s, dyn_lists=dyn_lists)
        if self.include_loans:
            self._build_loans(wb, sample=s)
        if self.include_pension:
            self._build_pension(wb, sample=s)
        if self.include_benefits:
            self._build_benefits(wb, sample=s)
        if self.include_disabilities:
            self._build_disabilities(wb, sample=s)
        if self.include_vacations:
            self._build_vacations(wb, sample=s)
        if self.include_overtime:
            self._build_overtime(wb, sample=s)
        if self.include_embargos:
            self._build_embargos(wb, sample=s)
        if self.include_bonos:
            self._build_bonos(wb, sample=s)
        if self.include_cobros:
            self._build_cobros_empleados(wb, sample=s)
        if self.include_acumulados:
            self._build_acumulados(wb, sample=s)
        if self.include_config:
            self._build_config_inicial(wb)

        self._build_catalogs(wb)

        # Serializar a bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        file_data = base64.b64encode(buf.read())

        # Guardar como attachment y devolver descarga
        company_slug = self.company_id.name.replace(' ', '_')[:20]
        filename     = f'Machote_Planilla_{company_slug}_v52858.xlsx'

        att = self.env['ir.attachment'].create({
            'name':     filename,
            'type':     'binary',
            'datas':    file_data,
            'res_model': self._name,
            'res_id':    self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        return {
            'type':   'ir.actions.act_url',
            'url':    f'/web/content/{att.id}?download=true',
            'target': 'self',
        }
