from odoo import models, api
import logging
import calendar
from datetime import date
from dateutil.relativedelta import relativedelta

_logger = logging.getLogger(__name__)


class PlanillaScheduledActions(models.AbstractModel):
    """
    Acciones programadas (cron jobs) del modulo Planilla CR.
    Se ejecutan automaticamente segun la frecuencia configurada.
    """
    _name = 'planilla.scheduled.actions'
    _description = 'Acciones Programadas Planilla CR'

    @api.model
    def cron_check_anniversaries(self):
        """
        Detecta empleados que cumplen aniversario laboral HOY y
        envia notificacion al correo de la empresa (RRHH).
        Corre: diariamente a las 7am.
        """
        today = date.today()
        template = self.env.ref(
            'planilla_cr.email_template_anniversary', raise_if_not_found=False
        )
        if not template:
            _logger.warning('Planilla CR: plantilla email_template_anniversary no encontrada.')
            return

        employees = self.env['hr.employee'].search([
            ('active', '=', True),
            ('entry_date', '!=', False),
        ])

        count = 0
        for emp in employees:
            # Cumple aniversario si hoy es el mismo mes/dia que su fecha de ingreso
            if (emp.entry_date.month == today.month and
                    emp.entry_date.day == today.day and
                    emp.entry_date.year < today.year):
                years = today.year - emp.entry_date.year
                _logger.info(
                    'Planilla CR: aniversario %s -- %s cumple %d ano(s)',
                    emp.name, today, years
                )
                try:
                    template.send_mail(emp.id, force_send=True)
                    count += 1
                except Exception as e:
                    _logger.error('Planilla CR: error enviando email aniversario %s: %s', emp.name, e)

        _logger.info('Planilla CR: cron_check_anniversaries -- %d notificaciones enviadas.', count)

    @api.model
    def cron_close_completed_loans(self):
        """
        Marca como pagados los prestamos donde todas las cuotas
        ya fueron descontadas. Corre: diariamente.
        """
        loans = self.env['planilla.employee.loan'].search([
            ('state', 'in', ('approved', 'active')),
        ])
        closed = 0
        for loan in loans:
            if loan.installment_ids and all(
                i.state in ('deducted', 'cancelled') for i in loan.installment_ids
            ):
                loan.action_check_paid()
                closed += 1
                _logger.info('Planilla CR: prestamo cerrado automaticamente -- %s / %s',
                             loan.employee_id.name, loan.name)


        _logger.info('Planilla CR: cron_close_completed_loans -- %d prestamos cerrados.', closed)

    @api.model
    def cron_alert_embargo_expiry(self):
        """
        Alerta cuando un embargo judicial esta proximo a vencer (30 dias antes de date_end).
        El patrono debe notificar al juzgado y al empleado antes del vencimiento.
        Art. 172 CT: responsabilidad solidaria del patrono.
        Corre: diariamente.
        """
        from datetime import date as _date
        from dateutil.relativedelta import relativedelta
        today = _date.today()
        threshold = today + relativedelta(days=30)

        embargos = self.env['planilla.embargo'].search([
            ('state', '=', 'active'),
            ('date_end', '!=', False),
            ('date_end', '>=', today),
            ('date_end', '<=', threshold),
        ])
        if not embargos:
            return

        # Agrupar por empresa para enviar un email por empresa
        by_company = {}
        for emb in embargos:
            comp = emb.company_id or self.env.company
            by_company.setdefault(comp, []).append(emb)

        for company, embs in by_company.items():
            if not company.email:
                continue
            lines = '\n'.join(
                f'   {e.employee_id.name}: expediente {e.numero_expediente} '
                f'-- vence {e.date_end} ({e.juzgado})'
                for e in embs
            )
            body = (
                f'<p><strong>Alerta de Embargos Judiciales por Vencer -- {company.name}</strong></p>'
                f'<p>Los siguientes <strong>{len(embs)}</strong> embargo(s) vencen '
                f'en los proximos 30 dias (Art. 172 CT):</p>'
                f'<pre style="background:#FFF3CD;padding:10px;">{lines}</pre>'
                f'<p>Contacte al juzgado correspondiente para renovar o finalizar el embargo.</p>'
            )
            try:
                self.env['mail.mail'].create({
                    'subject': f' Embargos por vencer -- {company.name}',
                    'email_to': company.email,
                    'body_html': body,
                    'auto_delete': True,
                }).send()
                _logger.info('Planilla CR: alerta embargo %d registros enviada a %s',
                             len(embs), company.name)
            except Exception as e:

                _logger.error('Planilla CR: error alerta embargo %s: %s', company.name, e)

    @api.model
    def cron_bono_antiguedad(self):
        """
        Crea automaticamente el bono de antiguedad para empleados que cumplen
        su aniversario laboral HOY, segun la tabla configurada en
        planilla.bono.antiguedad.config por empresa.
        
        En CR es comun que convenios colectivos o politicas de empresa
        reconozcan la antiguedad con un bono anual (porcentaje del salario
        o monto fijo segun tramo de anos).
        
        Corre: diariamente (misma frecuencia que cron_check_anniversaries).
        """
        today = date.today()
        BonoConfig = self.env['planilla.bono.antiguedad.config']
        BonoModel  = self.env['planilla.bono']

        employees = self.env['hr.employee'].search([
            ('active', '=', True),
            ('entry_date', '!=', False),
        ])

        created = 0
        for emp in employees:
            # Solo los que cumplen aniversario HOY y llevan >= 1 ano
            if not (emp.entry_date.month == today.month and
                    emp.entry_date.day == today.day and
                    emp.entry_date.year < today.year):
                continue

            years = today.year - emp.entry_date.year
            company_id = emp.company_id.id if emp.company_id else self.env.company.id

            # Buscar configuracion de antiguedad aplicable
            cfg = BonoConfig.get_config_for_years(company_id, years)
            if not cfg:
                continue  # No hay configuracion para este tramo -- no crear bono

            base_salary = emp.base_salary or 0.0
            if not base_salary:
                _logger.warning('Planilla CR: bono antiguedad -- %s no tiene salario base', emp.name)
                continue

            monto = cfg.compute_bono_amount(base_salary, years)
            if monto <= 0:
                continue

            # Verificar si ya existe un bono de antiguedad para este aniversario
            # (evitar duplicados si el cron corre dos veces)
            bono_name = f'Bono Antiguedad -- {years} ano(s)'
            year_start = today.replace(month=1, day=1)
            existing = BonoModel.search([
                ('employee_id', '=', emp.id),
                ('bono_type', '=', 'antiguedad'),
                ('name', '=', bono_name),
                ('date_start', '>=', year_start),
            ], limit=1)
            if existing:
                continue

            try:
                BonoModel.create({
                    'employee_id':   emp.id,
                    'name':          bono_name,
                    'bono_type':     'antiguedad',
                    'amount_type':   cfg.amount_type,
                    'amount':        monto if cfg.amount_type == 'fixed' else 0.0,
                    'percentage':    cfg.percentage if cfg.amount_type == 'percentage' else 0.0,
                    'is_recurring':  False,  # Es puntual: solo en el mes del aniversario
                    'afecto_ccss':   True,   # Antiguedad es salarial (Art. 162 CT)
                    'afecto_renta':  True,
                    'date_start':    today,
                    'date_end':      today.replace(
                                         day=calendar.monthrange(today.year, today.month)[1]),
                    'state':         'active',
                    'note':          f'Creado automaticamente por el sistema al cumplir {years} ano(s) de servicio.',
                })
                created += 1
                _logger.info('Planilla CR: bono antiguedad creado para %s -- %d ano(s), CRC%s',
                             emp.name, years, f'{monto:,.2f}')
            except Exception as e:
                _logger.error('Planilla CR: error creando bono antiguedad para %s: %s', emp.name, e)

        _logger.info('Planilla CR: cron_bono_antiguedad -- %d bonos creados.', created)


    @api.model
    def cron_recompute_vacation_balances(self):
        """
        Recomputa el saldo de vacaciones de todos los empleados activos.
        Necesario porque _compute_vacation_balance usa date.today() que
        cambia diariamente pero no es un campo de Odoo que dispare @api.depends.
        Corre: diariamente para que el saldo sea correcto cada dia laboral.
        Art. 153 CT: 12 dias habiles por cada 50 semanas laboradas.
        """
        employees = self.env['hr.employee'].search([('active', '=', True)])
        if not employees:
            return
        # Forzar recompute real de campos store=True
        # invalidate_recordset solo limpia cache Python, no dispara recompute en BD
        # _compute_vacation_balance() escribe directamente los valores calculados
        employees._compute_vacation_balance()
        employees.flush_recordset()
        _logger.info(
            'Planilla CR: cron_recompute_vacation_balances -- %d empleados actualizados.',
            len(employees)
        )


    @api.model
    def cron_extra_vacation_days_new_year(self):
        """
        Notifica el aniversario laboral de cada empleado (fecha de ingreso,
        no el 1 de enero) para que quede registro en el chatter.

        IMPORTANTE (fix doble conteo): este cron YA NO escribe
        vacation_initial_balance ni vacation_initial_balance_date.
        _compute_vacation_balance() ya aplica el bono de aniversario EN VIVO
        cada vez que se recalcula el saldo (su "Fase 3"), usando entry_date
        y esta misma configuracion -- no necesita que este cron le escriba
        nada. Antes este cron SI escribia el bono directo al saldo base, lo
        cual creaba una segunda fuente de verdad: en el primer aniversario
        de cada empleado coincidia por casualidad con el calculo en vivo
        (porque tambien fijaba la fecha de corte esa primera vez), pero del
        SEGUNDO aniversario en adelante (la fecha de corte ya no se volvia
        a mover) el calculo en vivo volvia a sumar el mismo bono que este
        cron ya habia escrito -- duplicando el saldo cada aniversario
        posterior al primero.

        Logica (Art. 153 CT):
        - Si hoy es el mismo mes y dia que la fecha de ingreso del empleado,
          y el empleado lleva al menos 1 ano completo, se notifica el
          aniversario y los dias que corresponden segun la configuracion.

        Modalidades:
        - 'fixed': todos reciben la misma cantidad de dias configurada.
        - 'per_year': N dias x anos completos de servicio (ej: 2d x 3 anos = 6d).

        Corre diariamente. Solo aplica a empleados cuyo aniversario es HOY.
        """
        from datetime import date as _date
        today = _date.today()

        configs = self.env['planilla.accounting.config'].search([
            ('extra_vacation_days_enabled', '=', True),
        ])

        for config in configs:
            base_days = config.extra_vacation_days_amount
            if base_days <= 0:
                continue

            employees = self.env['hr.employee'].search([
                ('company_id', '=', config.company_id.id),
                ('active', '=', True),
            ])

            applied = 0
            details = []

            for emp in employees:
                if not emp.entry_date:
                    continue

                # Verificar si HOY es el aniversario del empleado
                # (mismo mes y dia que su fecha de ingreso)
                if emp.entry_date.month != today.month or emp.entry_date.day != today.day:
                    continue

                # Anos completos de servicio hoy
                years_served = (today - emp.entry_date).days // 365
                if years_served <= 0:
                    # Menos de 1 ano: no aplica aun
                    continue

                # Evitar notificar dos veces el mismo aniversario si el cron
                # corre mas de una vez el mismo dia (no controla el saldo,
                # solo evita spam de notificaciones)
                last_applied = emp.vacation_last_anniversary_year or 0
                if last_applied >= today.year:
                    continue

                days_to_add = base_days

                # Marcar el ano notificado (unico proposito: evitar spam de
                # notificaciones repetidas -- NO controla el saldo)
                emp.write({'vacation_last_anniversary_year': today.year})

                # Registrar en chatter para trazabilidad. El saldo real ya
                # incluye este bono automaticamente via _compute_vacation_balance,
                # que completa hasta days_to_add (1 dia normal del ciclo +
                # el resto vía el bono de aniversario, no se suman encima).
                emp.message_post(
                    body=(
                        f'<b>Aniversario laboral {today.strftime("%d/%m/%Y")}:</b> '
                        f'{years_served} ano(s) de servicio. '
                        f'Total de <b>{days_to_add} dias</b> de vacaciones este mes '
                        f'(1 dia del ciclo normal + {max(days_to_add - 1, 0)} del bono '
                        f'de aniversario) -- ya reflejado automaticamente en el saldo.'
                    ),
                    message_type='notification',
                )

                applied += 1
                details.append(f'{emp.name}: +{days_to_add}d ({years_served} anos)')

            if applied:
                _logger.info(
                    'Planilla CR: cron_aniversario_vacaciones -- %s: %d empleados notificados. %s',
                    config.company_id.name, applied,
                    ' | '.join(details[:10]) + ('...' if len(details) > 10 else '')
                )

    def cron_alert_document_expiry(self):
        """
        Alerta cuando un documento personal de un empleado (cedula, licencia
        de conducir, carne de manipulacion de alimentos, permiso de trabajo,
        etc.) esta por vencer o ya vencio.

        Dos canales de aviso, independientes:
          1. Chatter del documento -- una sola vez por vencimiento (via el
             campo alert_sent, se reinicia solo si se corrige la fecha de
             vencimiento, ej. al renovar el documento).
          2. Correo con Excel adjunto a planilla.accounting.config.
             document_alert_emails -- se reenvia CADA DIA mientras sigan
             existiendo documentos por vencer/vencidos para esa empresa
             (mismo patron que cron_alert_embargo_expiry: es un recordatorio
             recurrente, no una notificacion de una sola vez, para que no se
             pierda de vista mientras el documento siga sin renovarse).

        Corre: diariamente.
        """
        from markupsafe import escape as _esc

        # -- Canal 1: chatter, solo documentos nuevos (alert_sent=False) ----
        docs_nuevos = self.env['planilla.employee.document'].search([
            ('state', 'in', ('por_vencer', 'vencido')),
            ('alert_sent', '=', False),
        ])
        for doc in docs_nuevos:
            _emp_name = _esc(doc.employee_id.name or '')
            _doc_name = _esc(doc.document_type_id.name or '')
            if doc.state == 'vencido':
                msg = (
                    f'<b>Documento vencido:</b> {_doc_name} '
                    f'de {_emp_name} vencio el '
                    f'{doc.expiry_date.strftime("%d/%m/%Y")} '
                    f'({abs(doc.days_to_expiry)} dias atras). Renovar cuanto antes.'
                )
            else:
                msg = (
                    f'<b>Documento por vencer:</b> {_doc_name} '
                    f'de {_emp_name} vence el '
                    f'{doc.expiry_date.strftime("%d/%m/%Y")} '
                    f'({doc.days_to_expiry} dias restantes).'
                )
            doc.message_post(body=msg, message_type='notification')
            doc.alert_sent = True

        if docs_nuevos:
            _logger.info(
                'Planilla CR: cron_alert_document_expiry -- %d documento(s) nuevos notificados en chatter.',
                len(docs_nuevos)
            )

        # -- Canal 2: correo con Excel, TODOS los que sigan pendientes ------
        todos_pendientes = self.env['planilla.employee.document'].search([
            ('state', 'in', ('por_vencer', 'vencido')),
        ])
        if not todos_pendientes:
            return

        by_company = {}
        for doc in todos_pendientes:
            comp = doc.company_id or self.env.company
            by_company.setdefault(comp, []).append(doc)

        for company, docs in by_company.items():
            cfg = self.env['planilla.accounting.config'].search(
                [('company_id', '=', company.id)], limit=1)
            emails_raw = cfg.document_alert_emails if cfg else False
            if not emails_raw:
                continue
            emails = [e.strip() for e in emails_raw.split(',') if e.strip()]
            if not emails:
                continue

            # Respetar la frecuencia configurada (diaria/semanal/mensual) --
            # el cron corre todos los dias, pero el correo consolidado solo
            # se manda cuando ya paso el intervalo configurado desde el
            # ultimo envio. El aviso en el chatter de cada documento (Canal 1,
            # arriba) sigue siendo inmediato sin importar esto.
            freq = cfg.document_alert_frequency or 'weekly'
            last_sent = cfg.document_alert_last_sent
            hoy = date.today()
            if last_sent:
                dias_desde_ultimo = (hoy - last_sent).days
                intervalo_minimo = {'daily': 1, 'weekly': 7, 'monthly': 30}.get(freq, 7)
                if dias_desde_ultimo < intervalo_minimo:
                    continue  # todavia no toca enviar para esta empresa

            xlsx_data = self._build_document_alert_excel(docs)
            n_vencidos = sum(1 for d in docs if d.state == 'vencido')
            n_por_vencer = len(docs) - n_vencidos
            freq_label = {'daily': 'diariamente', 'weekly': 'semanalmente',
                          'monthly': 'mensualmente'}.get(freq, 'periodicamente')

            mail_values = {
                'subject': f'[{company.name}] Documentos de empleado por vencer/vencidos '
                           f'({n_vencidos} vencidos, {n_por_vencer} por vencer)',
                'body_html': (
                    f'<p>Adjunto el detalle de documentos de empleado que estan '
                    f'<b>vencidos ({n_vencidos})</b> o <b>por vencer ({n_por_vencer})</b> '
                    f'en {_esc(company.name)}.</p>'
                    f'<p>Este correo se repite {freq_label} mientras sigan pendientes '
                    f'de renovar (frecuencia configurable en Configuracion Contable).</p>'
                ),
                'email_to': ','.join(emails),
                'email_from': company.email or self.env.user.email or '',
                'auto_delete': False,
                'attachment_ids': [(0, 0, {
                    'name': f'Documentos_por_vencer_{date.today()}.xlsx',
                    'datas': xlsx_data,
                    'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                })],
            }
            # Encolar sin enviar sincrono -- lo procesa el cron de correo de Odoo
            self.env['mail.mail'].create(mail_values)
            if cfg:
                cfg.document_alert_last_sent = hoy
            _logger.info(
                'Planilla CR: cron_alert_document_expiry -- correo con Excel enviado '
                'a %s para %s (%d vencidos, %d por vencer, frecuencia=%s).',
                emails, company.name, n_vencidos, n_por_vencer, freq
            )

    def _build_document_alert_excel(self, docs):
        """Genera el Excel de documentos por vencer/vencidos en memoria,
        retorna el contenido en base64 listo para adjuntar a un mail.mail."""
        import io
        import base64
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = 'Documentos por Vencer'

        headers = ['Empleado', 'Tipo de Documento', 'Numero', 'Fecha de Vencimiento',
                   'Dias Restantes', 'Estado']
        ws.append(headers)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        for col_idx, _h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        vencido_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        por_vencer_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

        docs_sorted = docs.sorted(key=lambda d: d.expiry_date or date.max)
        for row_idx, doc in enumerate(docs_sorted, start=2):
            ws.append([
                doc.employee_id.name or '',
                doc.document_type_id.name or '',
                doc.document_number or '',
                doc.expiry_date.strftime('%d/%m/%Y') if doc.expiry_date else '',
                doc.days_to_expiry,
                'Vencido' if doc.state == 'vencido' else 'Por Vencer',
            ])
            fill = vencido_fill if doc.state == 'vencido' else por_vencer_fill
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

        for col_idx, width in enumerate([28, 30, 16, 20, 14, 14], start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = width

        buf = io.BytesIO()
        wb.save(buf)
        return base64.b64encode(buf.getvalue())

    @api.model
    def cron_alert_negative_vacations(self):
        """
        Envia un resumen semanal a cada empresa sobre empleados
        con saldo de vacaciones negativo. Corre: lunes a las 8am.
        """
        companies = self.env['res.company'].search([])
        for company in companies:
            employees_neg = self.env['hr.employee'].search([
                ('company_id', '=', company.id),
                ('active', '=', True),
                ('vacation_balance_alert', '=', True),
            ])
            if not employees_neg:
                continue

            if not company.email:
                _logger.warning(
                    'Planilla CR: empresa %s no tiene email configurado para alertas.', company.name
                )
                continue

            # Construir resumen en texto
            lines = '\n'.join(
                f"   {e.name}: {e.vacation_days_available:.1f} dias "
                f"(acumulados: {e.vacation_days_accrued:.1f}, tomados: {e.vacation_days_taken:.1f})"
                for e in employees_neg
            )
            body = (
                f"<p>Estimado equipo de RRHH de <strong>{company.name}</strong>,</p>"
                f"<p>Los siguientes <strong>{len(employees_neg)}</strong> empleado(s) tienen "
                f"saldo de vacaciones <strong style='color:#E74C3C;'>negativo</strong>:</p>"
                f"<pre style='background:#FDE8E8;padding:10px;border-radius:4px;'>{lines}</pre>"
                f"<p>Por favor revise y tome las acciones necesarias segun el Art. 153 CT CR.</p>"
            )
            try:
                self.env['mail.mail'].create({
                    'subject': f'WARN Alerta: {len(employees_neg)} empleado(s) con vacaciones negativas -- {company.name}',
                    'email_to': company.email,
                    'body_html': body,
                    'auto_delete': True,
                }).send()
                _logger.info(
                    'Planilla CR: alerta vacaciones negativas enviada a %s (%d empleados)',
                    company.name, len(employees_neg)
                )
            except Exception as e:
                _logger.error('Planilla CR: error enviando alerta vacaciones %s: %s', company.name, e)

    @api.model
    def cron_alert_prescribing_vacations(self):
        """
        Alerta sobre empleados con vacaciones proximas a prescribir.
        Art. 156 CT CR: vacaciones prescriben a los 2 anos de ganadas.
        Corre: primer dia de cada mes.
        """
        today = date.today()
        threshold = today + relativedelta(months=2)  # Alerta con 2 meses de anticipacion

        companies = self.env['res.company'].search([])
        for company in companies:
            employees = self.env['hr.employee'].search([
                ('company_id', '=', company.id),
                ('active', '=', True),
                ('entry_date', '!=', False),
            ])
            # L2 FIX: batch query -- traer ultima vacacion de todos los empleados en 1 query
            emp_ids = employees.filtered(lambda e: e.vacation_days_available > 0).ids
            last_vacs = {}
            if emp_ids:
                vac_groups = self.env['planilla.vacation.payment'].read_group(
                    domain=[
                        ('employee_id', 'in', emp_ids),
                        ('state', 'in', ('approved', 'paid')),
                        ('vacation_type', '=', 'disfrutadas'),
                    ],
                    # FIX M-05 v51: el campo correcto es date_start, no date_from
                    # (date_from no existe en planilla.vacation.payment -> error silencioso)
                    fields=['employee_id', 'date_start:max'],
                    groupby=['employee_id'],
                )
                last_vacs = {
                    g['employee_id'][0]: g['date_start']
                    for g in vac_groups if g.get('date_start')
                }
            at_risk = []
            for emp in employees:
                if emp.vacation_days_available > 0:
                    ref_date = last_vacs.get(emp.id) or emp.entry_date
                    if ref_date:
                        months_since = (today - ref_date).days / 30
                        if months_since >= 22:
                            at_risk.append((emp.name, emp.vacation_days_available, round(months_since, 0)))

            if not at_risk or not company.email:
                continue

            lines = '\n'.join(
                f"   {name}: {days:.1f} dias disponibles (sin tomar hace ~{months:.0f} meses)"
                for name, days, months in at_risk
            )
            body = (
                f"<p>Alerta de prescripcion de vacaciones -- <strong>{company.name}</strong></p>"
                f"<p><strong>{len(at_risk)}</strong> empleado(s) tienen vacaciones en riesgo de prescribir "
                f"(Art. 156 CT CR -- prescriben a los 2 anos):</p>"
                f"<pre style='background:#FFF3CD;padding:10px;border-radius:4px;'>{lines}</pre>"
                f"<p>Se recomienda programar sus vacaciones antes de que prescriban.</p>"
            )
            try:
                self.env['mail.mail'].create({
                    'subject': f' Vacaciones proximas a prescribir -- {company.name}',
                    'email_to': company.email,
                    'body_html': body,
                    'auto_delete': True,
                }).send()
            except Exception as e:
                _logger.error('Planilla CR: error alerta prescripcion %s: %s', company.name, e)
